from pymongo import MongoClient
import pandas as pd
import dns.resolver
import certifi
from bson import ObjectId, errors
import config
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

# DNS resolver and MongoDB connection setup
dns.resolver.default_resolver = dns.resolver.Resolver(configure=False)
dns.resolver.default_resolver.nameservers = ['8.8.8.8']

carrier_list = ["AF", "KL","DL", "MU", "TO", "HV"]

client = MongoClient(f'mongodb+srv://{config.mongo_pat}', tlsCAFile=certifi.where())

# MongoDB collections
db = client['legacy-api-management']
col_items = db["items"]
col_bills = db["bills"]
col_users = db["users"]
col_soc = db["societies"]
col_billings = db["billings"]

airports_df = pd.read_csv("const/Air Airports and cities codes - O&D  2024.csv", delimiter=";")
rail_metro_df = pd.read_csv("const/O&D Metropolis rail 2024 v In&Out.csv",delimiter=";")
rail_euro_df = pd.read_csv("const/O&D European Rail 2024 v In&Out.csv",delimiter=";")

current_date = datetime.now()
current_year = current_date.year
last_month = current_date.month - 1 if current_date.month > 1 else 12
last_month_name = datetime(1900, last_month, 1).strftime('%B')
if last_month == 12:
    current_year -= 1  # Adjust the year if the last month is December from the previous year

def get_soc(id_soc):
    result = col_soc.find_one({"_id": ObjectId(f"{id_soc}")},
                              {"name": 1, "settings.flight.bluebizz": 1, "_id": 0})

    name = result.get("name", None)
    bluebizz = result.get("settings", {}).get("flight", {}).get("bluebizz", None)
    bluebizz_value = bluebizz if bluebizz else "Bluebizz"

    osi = bluebizz_value
    name_orga = name

    return osi,name_orga

def get_bills(id_soc,start_date):
    import pandas as pd
    from pymongo import MongoClient

    # Définir les critères de filtrage
    filter_criteria = {
        "societyId": f"{id_soc}",
        "createdAt": {"$gte": start_date},
        "type": {"$in": ["receipt", "credit", "unitary"]}
    }

    # Requête MongoDB pour extraire les informations des lignes
    projection = {
        "type": 1,  # Projeter 'type' à la racine
        "lines.itemId": 1,
        "lines.label": 1,
        "lines.type": 1,
        "lines.price.amount": 1,
        "lines.userIds": 1,
        "_id": 0  # Exclure le champ _id des résultats
    }

    # Effectuer la requête
    documents = col_bills.find(filter_criteria, projection)

    # Liste pour stocker les résultats
    data = []

    # Extraire les résultats et aplatir les valeurs des lignes
    for doc in documents:
        doc_type = doc.get("type")  # Récupérer le 'type' à la racine
        for line in doc.get('lines', []):
            # Extraire les valeurs des sous-documents
            item_data = {
                "ITEM_ID": line.get("itemId"),
                "type_bill": doc_type,  # Associer le type du document à chaque ligne
                "label": line.get("label"),
                "type": line.get("type"),
                "TOTAL_BILLED": line.get("price", {}).get("amount"),
                "userIds": ', '.join([str(user_id) for user_id in line.get("userIds", [])])
            }
            data.append(item_data)

    # Convertir la liste en DataFrame
    df = pd.DataFrame(data)

    df.to_csv(f"csv/base/bills_{id_soc}.csv")
    # Afficher le DataFrame
    print(df)

################ 🛫FLIGHT 🛫 #####################

def fetch_flight_details(item_id):
    # Chercher l'item dans la collection MongoDB
    item = col_items.find_one({"id": item_id})
    print(item_id)
    if item and item.get('type') == 'flight':
        trips = item.get('detail', {}).get('trips', [])
        all_legs_details = []

        # Boucle sur tous les trips
        for trip_index, trip in enumerate(trips):
            legs = trip.get('legs', [])

            if legs:
                # Si 'legs' existe, traiter chaque leg
                for leg_index, leg in enumerate(legs):
                    leg_details = {
                        f"{leg_index}_ori_city": leg.get('departure', {}).get('city', ''),
                        f"{leg_index}_des_city": leg.get('arrival', {}).get('city', ''),
                        f"{leg_index}_ori_country": leg.get('departure', {}).get('country', ''),
                        f"{leg_index}_des_country": leg.get('arrival', {}).get('country', ''),
                        f"{leg_index}_governing_carrier": leg.get('governingCarrier', '')
                    }
                    all_legs_details.append(leg_details)
            else:
                # Si 'legs' n'existe pas, utiliser les chemins alternatifs
                leg_details = {
                    f"{trip_index}_ori_city": trip.get('Dep', {}).get('IATALocationCode', ''),
                    f"{trip_index}_des_city": trip.get('Arrival', {}).get('IATALocationCode', ''),
                    f"{trip_index}_ori_country": "NDC",  # Valeur par défaut
                    f"{trip_index}_des_country": "NDC",  # Valeur par défaut
                    f"{trip_index}_governing_carrier": trip.get('OperatingCarrierInfo', {}).get('CarrierDesigCode', '')
                }
                all_legs_details.append(leg_details)

        # Fusionner tous les détails des legs dans un seul dictionnaire
        combined_leg_details = {}
        for leg_detail in all_legs_details:
            combined_leg_details.update(leg_detail)

        return combined_leg_details
    return {}

def concatenate_cities(row):
    number_of_legs = row['NB_LEGS']

    # Si le nombre de legs est supérieur à 3, renvoyer "check"
    if number_of_legs > 3:
        return "check"

    # Si le nombre de legs est inférieur ou égal à 2
    ori_city = row.get('0_ori_city', '')
    des_city = row.get('0_des_city', '')
    ori_country = row.get('0_ori_country', '')
    des_country = row.get('0_des_country', '')

    # Si 'NDC' est dans '0_ori_country', on change les variables pour utiliser les codes NDC
    if ori_country == 'NDC':
        ori_city = row.get('0_Ori_NDC', '')
        des_city = row.get('0_Des_NDC', '')
        ori_country = row.get('0_Ori_country_NDC', '')
        des_country = row.get('0_Des_country_NDC', '')

    special_cities = ['PAR', 'BIO', 'BRU', 'FRA', 'GVA', 'LUX']

    # Gérer le cas où PAR est dans ori_city ou des_city
    if 'PAR' in [ori_city, des_city]:
        if ori_city == 'PAR':
            ori_city, des_city = 'PAR', des_city
        elif des_city == 'PAR':
            ori_city, des_city = 'PAR', ori_city

        if number_of_legs <= 2:
            return f"{ori_city} {des_city}"
        else:
            return f"{ori_city} {des_city} !!!"

    # Gérer les autres villes spéciales (BIO, BRU, FRA, GVA, LUX)
    elif ori_city in special_cities or des_city in special_cities:
        # Trier par ordre alphabétique si une des villes spéciales est présente
        sorted_cities = sorted([ori_city, des_city])
        if number_of_legs <= 2:
            return f"{sorted_cities[0]} {sorted_cities[1]}"
        else:
            return f"{sorted_cities[0]} {sorted_cities[1]} !!!"

    # Gérer le cas où l'origine et la destination sont en France
    elif ori_country == 'FR' and des_country == 'FR':
        sorted_cities = sorted([ori_city, des_city])
        if number_of_legs <= 2:
            return f"{sorted_cities[0]} {sorted_cities[1]}"
        else:
            return f"{sorted_cities[0]} {sorted_cities[1]} !!!"

    # Gérer le cas où l'origine et la destination ne sont pas en France
    elif ori_country != 'FR' and des_country != 'FR':
        sorted_cities = sorted([ori_city, des_city])
        if number_of_legs <= 2:
            return f"{sorted_cities[0]} {sorted_cities[1]}"
        else:
            return f"{sorted_cities[0]} {sorted_cities[1]} !!!"

    # Gérer le cas où l'origine est en France et la destination est à l'étranger
    elif ori_country == 'FR' and des_country != 'FR':
        if number_of_legs <= 2:
            return f"{ori_city} {des_city}"
        else:
            return f"{ori_city} {des_city} !!!"

    # Gérer le cas où l'origine est à l'étranger et la destination est en France
    elif ori_country != 'FR' and des_country == 'FR':
        if number_of_legs <= 2:
            return f"{des_city} {ori_city}"
        else:
            return f"{des_city} {ori_city} !!!"

    return ""

def concatenate_carriers(row):
    carrier_columns = [col for col in row.index if col.endswith('_governing_carrier')]
    return ' '.join([str(row[col]) for col in carrier_columns if row[col]])

def classify_airline(concat_carrier):
    if any(carrier in concat_carrier for carrier in carrier_list):
        return "AIR FRANCE"
    return "INDUSTRIE"

def lookup_airport_details(concatenated_city):
    # Find the matching row where 'O&D restitué' matches the given concatenated_city
    match_row = airports_df[airports_df['O&D restitué'] == concatenated_city]

    # If no match found, try to find in 'O&D Inbound'
    if match_row.empty:
        match_row = airports_df[airports_df['O&D Inbound'] == concatenated_city]

    if not match_row.empty:
        # Extract values for each field if a match is found
        haul_type = match_row['Haul type'].values[0]
        origin_area = match_row['Origin area'].values[0]
        annex_c = match_row['Annex C/ Out of Annex C 2023'].values[0]
        label_origin = match_row['Label cities of origin'].values[0]
        label_destination = match_row['Label cities of destination'].values[0]
        od_restitue = match_row['O&D restitué'].values[0]  # Add the 'O&D restitué' column value

        # Return all the extracted values
        return haul_type, origin_area, annex_c, label_origin, label_destination, od_restitue

    # Return empty strings if no match is found
    return "", "", "", "", "", ""

def lookup_ndc_code(city, primary_column, primary_code_column, secondary_column=None, secondary_code_column=None):
    # Chercher la ville dans le fichier CSV dans la colonne primaire (e.g., 'Airport origin code')
    code_row = airports_df[airports_df[primary_column] == city]

    if not code_row.empty:
        return code_row[primary_code_column].values[0]

    # Si aucune correspondance trouvée et les colonnes secondaires sont spécifiées
    if secondary_column and secondary_code_column:
        code_row = airports_df[airports_df[secondary_column] == city]
        if not code_row.empty:
            return code_row[secondary_code_column].values[0]

    return ""

def get_items_flight(id_soc, start_date):
    filter_criteria = {
        "society._id": ObjectId(id_soc),
        "createdAt": {"$gte": start_date},
        "type": "flight",
        "statusHistory.to": "confirmed",
        "status": {"$in": ["confirmed", "cancelled"]}
    }

    projection = {
        "id": 1,
        "offline": 1,
        "createdAt": 1,
        "type": 1,
        "status": 1,
        "billingId": 1,
        "travelers.billingId": 1,
        "detail.trips": 1,
        "_id": 0
    }

    items = col_items.find(filter_criteria, projection)
    data = []

    for item in items:
        offline_status = "Offline" if item.get("offline") else "Online"
        item_type = item.get("type")
        item_createdAt = item.get("createdAt")
        item_status = item.get("status", "Unknown")

        billing_id = item.get("billingId")
        if not billing_id:
            travelers = item.get("travelers", [])
            billing_id = next((traveler.get("billingId") for traveler in travelers if traveler.get("billingId")), None)

        raison = "Unknown"
        if billing_id:
            try:
                billing_doc = col_billings.find_one({"_id": ObjectId(billing_id)}, {"raison": 1})
            except errors.InvalidId:
                billing_doc = col_billings.find_one({"_id": billing_id}, {"raison": 1})

            if billing_doc:
                raison = billing_doc.get("raison", "Unknown")

        trips = item.get("detail", {}).get("trips", [])
        number_of_legs = sum([len(trip.get("legs", [])) if trip.get("legs", []) else 1 for trip in trips])

        basic_data = {
            "ITEM_ID": item.get("id"),
            "IS_OFFLINE": offline_status,
            "CREATED_AT": item_createdAt,
            "TYPE": item_type,
            "STATUS": item_status,
            "NB_LEGS": number_of_legs,
            "BILLING_ID": billing_id,
            "ENTITY": raison
        }

        leg_details = fetch_flight_details(item.get("id"))
        combined_data = {**basic_data, **leg_details}
        data.append(combined_data)

    df = pd.DataFrame(data)

    # Interroger le fichier CSV pour les colonnes avec 'NDC' dans les pays
    for i in range(2):  # Pour les index 0 et 1 (0_ori_country, 1_ori_country, etc.)
        ori_country_col = f'{i}_ori_country'
        des_country_col = f'{i}_des_country'
        ori_city_col = f'{i}_ori_city'
        des_city_col = f'{i}_des_city'

        # Si "NDC" dans ori_country, chercher dans 'Airport origin code', sinon dans 'Airport destination code'
        df[f'{i}_Ori_NDC'] = df.apply(
            lambda row: lookup_ndc_code(row[ori_city_col],
                                        'Airport origin code',
                                        'Origin Code',
                                        'Airport destination code',
                                        'Destination Code') if row[ori_country_col] == 'NDC' else '', axis=1)

        # Si "NDC" dans des_country, chercher dans 'Airport destination code', sinon dans 'Airport origin code'
        df[f'{i}_Des_NDC'] = df.apply(
            lambda row: lookup_ndc_code(row[des_city_col],
                                        'Airport destination code',
                                        'Destination Code',
                                        'Airport origin code',
                                        'Origin Code') if row[des_country_col] == 'NDC' else '', axis=1)

        # Condition pour Ori_country_NDC (Nouveau code)
        df[f'{i}_Ori_country_NDC'] = df.apply(
            lambda row: lookup_ndc_code(row[ori_city_col],
                                        'Airport origin code',
                                        'Country code of Origine',
                                        'Airport destination code',
                                        'Country code of destination') if row[ori_country_col] == 'NDC' else '', axis=1)

        # Condition pour Des_country_NDC (Nouveau code)
        df[f'{i}_Des_country_NDC'] = df.apply(
            lambda row: lookup_ndc_code(row[des_city_col],
                                        'Airport destination code',
                                        'Country code of destination',
                                        'Airport origin code',
                                        'Country code of Origine') if row[des_country_col] == 'NDC' else '', axis=1)


    # Concaténer les colonnes qui terminent par 'governing_carrier' pour créer CONCAT_carrier
    df['CONCAT_carrier'] = df.apply(concatenate_carriers, axis=1)

    # Créer la colonne 'AIRLINE_GROUP' pour classifier en fonction de la présence d'un code de 'carrier_list'
    df['AIRLINE_GROUP'] = df['CONCAT_carrier'].apply(classify_airline)

    # Sauvegarder le DataFrame en CSV
    df.to_csv(f"csv/base/items_flight_{id_soc}.csv", index=False)

def merge_extract_flight(id_soc):
    # Charger les fichiers CSV de vol et fusionner
    df0 = pd.read_csv(f"csv/base/bills_{id_soc}.csv")
    df0 = df0[df0["type"] == "flight"]
    df1 = pd.read_csv(f"csv/base/items_flight_{id_soc}.csv")
    df = pd.merge(df0, df1, on='ITEM_ID', how='inner')

    # Convertir CREATED_AT en formats YYYY-MM et YYYY
    df['ISSUED_MONTH'] = pd.to_datetime(df['CREATED_AT']).dt.strftime('%Y%m')
    df['ODLIST'] = pd.to_datetime(df['CREATED_AT']).dt.strftime('%Y')

    # Appliquer la fonction concatenate_cities à chaque ligne du DataFrame
    df['CONCATENATED_CITIES'] = df.apply(concatenate_cities, axis=1)

    # Rechercher les détails supplémentaires dans 'airports_df' pour 'CONCATENATED_CITIES'
    df[['HAUL_TYPE', 'ORIGIN_AREA', 'ANNEX_C', 'LABEL_ORIGIN', 'LABEL_DESTINATION','O&D RESTITUE']] = df.apply(
        lambda row: pd.Series(lookup_airport_details(row['CONCATENATED_CITIES'])), axis=1)

    columns_to_keep = [
        'ITEM_ID', 'type_bill', 'ISSUED_MONTH', 'ODLIST', 'IS_OFFLINE', 'TOTAL_BILLED',
        'NB_LEGS', 'CONCATENATED_CITIES', 'HAUL_TYPE', 'ORIGIN_AREA', 'ANNEX_C',
        'LABEL_ORIGIN', 'LABEL_DESTINATION', 'AIRLINE_GROUP','O&D RESTITUE'
    ]

    # Sélectionner uniquement les colonnes spécifiées
    df_filtered = df[columns_to_keep]

    # Trier le DataFrame par 'ODLIST'
    df_filtered = df_filtered.sort_values(by='ODLIST')

    # Sauvegarder le DataFrame filtré dans un fichier CSV
    df_filtered.to_csv(f'csv/res/flight/raw_flight_{id_soc}.csv', index=False)

    df_filtered = df_filtered[df_filtered['O&D RESTITUE'] != ""]
    df_filtered.to_csv(f'csv/res/flight/filtered_flight_{id_soc}.csv', index=False)

def group_flight(id_soc,iata,osi):
    # Charger le fichier CSV téléchargé
    file_path = f'csv/res/flight/filtered_flight_{id_soc}.csv'
    df = pd.read_csv(file_path)

    # Groupement par les colonnes spécifiées
    grouping_columns = [
        'O&D RESTITUE', 'ORIGIN_AREA', 'ANNEX_C', 'LABEL_ORIGIN',
        'LABEL_DESTINATION', 'ISSUED_MONTH', 'ODLIST', 'IS_OFFLINE'
    ]

    # Agrégation des données
    df_grouped_af = df.groupby(grouping_columns).agg(
        TOTAL_BILLED_AIR_FRANCE=('TOTAL_BILLED', lambda x: x[df['AIRLINE_GROUP'] == 'AIR FRANCE'].sum()),
        TOTAL_BILLED_INDUSTRIE=('TOTAL_BILLED', lambda x: x[(df['AIRLINE_GROUP'] == 'AIR FRANCE') | (df['AIRLINE_GROUP'] == 'INDUSTRIE')].sum()),
        NB_LEGS_AIR_FRANCE=('NB_LEGS', lambda x: x[df['AIRLINE_GROUP'] == 'AIR FRANCE'].sum()),
        NB_LEGS_INDUSTRIE=('NB_LEGS', lambda x: x[(df['AIRLINE_GROUP'] == 'AIR FRANCE') | (df['AIRLINE_GROUP'] == 'INDUSTRIE')].sum())
    ).reset_index()


    # fichier brut
    df_grouped = df.groupby(grouping_columns).agg({
        'TOTAL_BILLED': 'sum',
        'NB_LEGS': 'sum'}).reset_index()

    df_grouped.to_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv")

    # fichier af

    df_grouped_af[['CODE ORIGINE', 'CODE DESTINATION']] = df_grouped_af['O&D RESTITUE'].str.split(' ', n=1, expand=True)

    # Renommer les colonnes
    df_grouped_af = df_grouped_af.rename(columns={
        "O&D RESTITUE" : "O&D",
        "ISSUED_MONTH": "DATE D'EMISSION",
        'ORIGIN_AREA': 'ZONE ORIGINE',
        'ANNEX_C': 'PERIMETRE',
        'LABEL_ORIGIN': 'LIBELLE ORIGINE',
        'LABEL_DESTINATION': 'LIBELLE DESTINATION',
        'IS_OFFLINE': 'TYPE DE VENTE',
        'TOTAL_BILLED_AIR_FRANCE': 'CA GROUPE AF KL',
        'NB_LEGS_AIR_FRANCE': 'NB O&D GROUPE AF KL',
        'TOTAL_BILLED_INDUSTRIE': 'CA TOTAL INDUSTRIE',
        'NB_LEGS_INDUSTRIE': 'NB O&D INDUSTRIE',
    })
    df_grouped_af['CA GROUPE AF KL'] = df_grouped_af['CA GROUPE AF KL'].round(0).astype(int)
    df_grouped_af['NB O&D GROUPE AF KL'] = df_grouped_af['NB O&D GROUPE AF KL'].round(0).astype(int)
    df_grouped_af['CA TOTAL INDUSTRIE'] = df_grouped_af['CA TOTAL INDUSTRIE'].round(0).astype(int)
    df_grouped_af['NB O&D INDUSTRIE'] = df_grouped_af['NB O&D INDUSTRIE'].round(0).astype(int)
    df_grouped_af['REFERENCE AF'] = osi
    df_grouped_af['RAISON SOCIALE'] = ''
    df_grouped_af['IATA EMETTEUR'] = iata

    # Connexion en base pour recup OIN et NAME
    result = col_soc.find_one({"_id": ObjectId(id_soc)})

    # Extraire les variables "name" et "settings.flight.bluebizz"
    name = result.get("name", "Unknown")  # Assigner une valeur par défaut si name est None
    bluebizz = result.get("settings", {}).get("flight", {}).get("bluebizz", "Bluebizz")  # Valeur par défaut
    bluebizz_value = bluebizz if bluebizz else "Bluebizz"

    print(name, bluebizz_value)

    # Remplir les colonnes du DataFrame avec les valeurs récupérées
    df_grouped_af['REFERENCE AF'] = bluebizz_value
    df_grouped_af['RAISON SOCIALE'] = name

    # Trier les valeurs par "DATE D'EMISSION" et "O&D"
    df_grouped_af = df_grouped_af.sort_values(by=["DATE D'EMISSION", "O&D"], ascending=[False, True])

    # Vérifier si les valeurs de 'O&D' de flight_df sont présentes dans 'O&D restitué' de airports_df
    df_grouped_af['O&D_match'] = df_grouped_af['O&D'].isin(airports_df['O&D restitué'])

    # Filtrer les lignes où 'O&D' a une correspondance dans airports_df
    filtered_grouped_af = df_grouped_af[df_grouped_af['O&D_match']].drop(columns=['O&D_match'])

    # Sauvegarder le DataFrame filtré dans un nouveau fichier CSV
    filtered_grouped_af.to_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv", index=False)

    print(f"File saved: csv/res/flight/grouped_flight_{id_soc}.csv")

################ 🚅 TRAIN 🚅 #####################

def fetch_train_details(item_id):
    # Chercher l'item dans la collection MongoDB
    item = col_items.find_one({"id": item_id})
    print(item_id)

    if item and item.get('type') == 'train':
        journeys = item.get('detail', {}).get('journeys', [])
        nombre_leg = len(journeys)  # Nombre de voyages/jambes

        all_legs_details = []

        # Boucle sur tous les voyages (journeys)
        for journey_index, journey in enumerate(journeys):
            if journey.get('departure') and journey.get('arrival'):
                ori_locationId = journey['departure'].get('locationId', '')
                des_locationId = journey['arrival'].get('locationId', '')

                leg_details = {
                    f"{journey_index}_ori_name": journey['departure'].get('name', ''),
                    f"{journey_index}_des_name": journey['arrival'].get('name', ''),
                    f"{journey_index}_ori_locationId": ori_locationId,
                    f"{journey_index}_des_locationId": des_locationId,
                }
                all_legs_details.append(leg_details)
            else:
                # Si 'departure' et 'arrival' n'existent pas, chercher dans les 'segments'
                segments = journey.get('segments', [])
                for segment_index, segment in enumerate(segments):
                    ori_locationId = segment['departure'].get('locationId', '')
                    des_locationId = segment['arrival'].get('locationId', '')

                    leg_details = {
                        f"{segment_index}_ori_name": segment['departure'].get('name', ''),
                        f"{segment_index}_des_name": segment['arrival'].get('name', ''),
                        f"{segment_index}_ori_locationId": ori_locationId,
                        f"{segment_index}_des_locationId": des_locationId,
                    }
                    all_legs_details.append(leg_details)

        # Fusionner tous les détails des segments dans un seul dictionnaire
        combined_leg_details = {}
        for leg_detail in all_legs_details:
            combined_leg_details.update(leg_detail)

        return combined_leg_details, nombre_leg

    return {}, 0

def get_items_train(id_soc, start_date):
    filter_criteria = {
        "society._id": ObjectId(id_soc),
        "createdAt": {"$gte": start_date},
        "type": "train",
        "statusHistory.to": "confirmed",
        "status": {"$in": ["confirmed", "cancelled"]}
    }

    projection = {
        "id": 1,
        "offline": 1,
        "createdAt": 1,
        "type": 1,
        "status": 1,
        "billingId": 1,
        "travelers.billingId": 1,
        "detail.trips": 1,
        "_id": 0
    }

    items = col_items.find(filter_criteria, projection)
    data = []

    for item in items:
        offline_status = "Offline" if item.get("offline") else "Online"
        item_type = item.get("type")
        item_createdAt = item.get("createdAt")
        item_status = item.get("status", "Unknown")

        billing_id = item.get("billingId")
        if not billing_id:
            travelers = item.get("travelers", [])
            billing_id = next((traveler.get("billingId") for traveler in travelers if traveler.get("billingId")), None)

        raison = "Unknown"
        if billing_id:
            try:
                billing_doc = col_billings.find_one({"_id": ObjectId(billing_id)}, {"raison": 1})
            except errors.InvalidId:
                billing_doc = col_billings.find_one({"_id": billing_id}, {"raison": 1})

            if billing_doc:
                raison = billing_doc.get("raison", "Unknown")

        # Récupérer les détails des segments et le nombre de jambes (legs) avec `fetch_train_details`
        leg_details, nombre_leg = fetch_train_details(item.get("id"))

        # Initialisation de `basic_data`
        basic_data = {
            "ITEM_ID": item.get("id"),
            "IS_OFFLINE": offline_status,
            "CREATED_AT": item_createdAt,
            "TYPE": item_type,
            "STATUS": item_status,
            "NB_LEGS": nombre_leg,  # Utilisation de nombre_leg ici
            "BILLING_ID": billing_id,
            "ENTITY": raison
        }

        # Fusionner les données de base avec les détails des jambes
        combined_data = {**basic_data, **leg_details}

        # Ajout au DataFrame
        data.append(combined_data)

    # Créer le DataFrame
    df = pd.DataFrame(data)

    # Sauvegarder le DataFrame en CSV
    df.to_csv(f"csv/base/items_train_{id_soc}.csv", index=False)

    print(df.head())

def concat_and_remove_columns(id_soc):
    df = pd.read_csv(f"csv/base/items_train_{id_soc}.csv")
    ori_columns = [col for col in df.columns if col.startswith(tuple(f'{i}_ori_name' for i in range(2, 10)))]
    des_columns = [col for col in df.columns if col.startswith(tuple(f'{i}_des_name' for i in range(2, 10)))]
    ori_locationId_columns = [col for col in df.columns if col.startswith(tuple(f'{i}_ori_locationId' for i in range(2, 10)))]
    des_locationId_columns = [col for col in df.columns if col.startswith(tuple(f'{i}_des_locationId' for i in range(2, 10)))]

    # Concaténer les colonnes ori, des, locationId
    df['ori_concat'] = df[ori_columns].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
    df['des_concat'] = df[des_columns].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
    df['ori_locationId_concat'] = df[ori_locationId_columns].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
    df['des_locationId_concat'] = df[des_locationId_columns].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)

    # Supprimer les colonnes inutiles
    columns_to_remove = ori_columns + des_columns + ori_locationId_columns+ des_locationId_columns
    df.drop(columns=columns_to_remove, inplace=True)
    df.to_csv(f"csv/res/train/temp_{id_soc}.csv", index=False)

def normalize_station_names(id_soc):
    df = pd.read_csv(f"csv/res/train/temp_{id_soc}.csv")
    # Liste des colonnes à traiter
    name_columns = [col for col in df.columns if '_ori_name' in col or '_des_name' in col]

    # Dictionnaire des remplacements
    replacements = {
        "PARIS": "PARIS",
        "MONTPELLIER": "MONTPELLIER",
        "MARSEILLE": "MARSEILLE ST CHARLES",
        "BORDEAUX": "BORDEAUX",
        "STRASBOURG": "STRASBOURG",
        "LYON": "LYON"
    }

    # Fonction pour normaliser une entrée
    def normalize_name(name):
        if pd.isna(name):
            return name
        name_upper = name.upper()  # Mettre tout en majuscule pour ignorer la casse
        for keyword, replacement in replacements.items():
            if keyword in name_upper:
                return replacement
        return name  # Retourner le nom original s'il n'y a pas de correspondance

    # Appliquer la normalisation à toutes les colonnes pertinentes
    for col in name_columns:
        df[col] = df[col].apply(normalize_name)

    df.to_csv(f"csv/res/train/norm_{id_soc}.csv", index=False)

def fill_missing_location_ids(id_soc):
    # Charger les fichiers CSV
    temp_df = pd.read_csv(f"csv/res/train/norm_{id_soc}.csv")
    metropolis_df = pd.read_csv("const/O&D Metropolis rail 2024 v In&Out.csv", delimiter=";")
    european_df = pd.read_csv("const/O&D European Rail 2024 v In&Out.csv", delimiter=";")

    # Liste des colonnes à traiter
    location_columns = [
        '0_ori_locationId', '0_des_locationId', '1_ori_locationId', '1_des_locationId',
    ]

    # Liste des colonnes correspondantes pour les noms (origine et destination)
    name_columns = [
        '0_ori_name', '0_des_name', '1_ori_name', '1_des_name',
    ]

    # Fonction pour interroger les fichiers et remplir le locationId
    def lookup_location_id(name, ref_df):
        # Chercher dans 'Label of railway station of origin'
        origin_match = ref_df[ref_df['Label of railway station of origin'] == name]
        if not origin_match.empty:
            return origin_match['Railway station code NTS of origin'].values[0]
        else:
            # Chercher dans 'Label of railway station of destination'
            destination_match = ref_df[ref_df['Label of railway station of destination'] == name]
            if not destination_match.empty:
                return destination_match['Railway station code NTS of destination'].values[0]
        return "PAS TROUVÉ"

    # Remplir les valeurs manquantes pour chaque paire (locationId, name)
    for loc_col, name_col in zip(location_columns, name_columns):
        for idx, row in temp_df.iterrows():
            if pd.isna(row[loc_col]) and pd.notna(row[name_col]):
                # Chercher dans le fichier Metropolis
                result = lookup_location_id(row[name_col], metropolis_df)
                if result == "PAS TROUVÉ":
                    # Si pas trouvé dans Metropolis, chercher dans le fichier European
                    result = lookup_location_id(row[name_col], european_df)
                temp_df.at[idx, loc_col] = result
    temp_df = temp_df[(temp_df['0_ori_locationId'] != "PAS TROUVÉ") & (temp_df['0_des_locationId'] != "PAS TROUVÉ")]
    df = temp_df
    def determine_zone(row):
        # Vérifier si une des valeurs ne commence pas par 'FR'
        columns_to_check = ['0_ori_locationId', '0_des_locationId', '1_ori_locationId', '1_des_locationId']

        for col in columns_to_check:
            if pd.notna(row[col]) and not row[col].startswith("FR"):
                return "EUROPE"
        return "METRO"
    df['Zone'] = temp_df.apply(determine_zone, axis=1)

    # Créer la colonne 'Zone' en appliquant la fonction determine_zone à chaque ligne

    df.to_csv(f"csv/res/train/concat_{id_soc}.csv", index=False)

def process_locations(id_soc):
    # Charger le DataFrame
    df = pd.read_csv(f"csv/res/train/concat_{id_soc}.csv")
    rail_metro_df = pd.read_csv('const/O&D Metropolis rail 2024 v In&Out.csv', delimiter=";")
    rail_euro_df = pd.read_csv('const/O&D European Rail 2024 v In&Out.csv', delimiter=";")

    # Fonction pour déterminer les 'ori_location_rest' et 'des_location_rest'
    def determine_rest_locations(row):
        if row['0_ori_name'] == 'PARIS':
            return row['0_ori_locationId'], row['0_des_locationId']
        elif row['0_des_name'] == 'PARIS':
            return row['0_des_locationId'], row['0_ori_locationId']
        else:
            locations = sorted([row['0_ori_locationId'], row['0_des_locationId']])
            ori_name, des_name = sorted([row['0_ori_name'], row['0_des_name']])
            if ori_name == row['0_des_name']:
                return row['0_des_locationId'], row['0_ori_locationId']
            return locations[0], locations[1]

    # Créer les colonnes 'ori_location_rest' et 'des_location_rest'
    df['ori_location_rest'], df['des_location_rest'] = zip(*df.apply(determine_rest_locations, axis=1))

    # Fonction pour interroger rail_metro_df si la zone est 'METRO'
    def query_metropolis(ori_location, des_location):
        ori_code, ori_area, ori_label, des_code, des_label, des_country = None, None, None, None, None, None

        ori_match = rail_metro_df[rail_metro_df['Railway station code NTS of origin'] == ori_location]
        if not ori_match.empty:
            ori_code = ori_match['Origin code'].values[0]
            ori_area = ori_match['Origin area'].values[0]
            ori_label = ori_match['Label city of origin '].values[0]

        des_match = rail_metro_df[rail_metro_df['Railway station code NTS of destination'] == des_location]
        if not des_match.empty:
            des_code = des_match['Destination code'].values[0]
            des_country = des_match['Country code of destination'].values[0]
            des_label = des_match['Label city of destination '].values[0]

        return ori_code, ori_label, ori_area, des_code, des_label, des_country

    # Fonction pour interroger rail_euro_df si la zone est 'EUROPE'
    def query_europe(ori_location, des_location):
        ori_code, ori_area, ori_label, des_code, des_label, des_country = None, None, None, None, None, None

        ori_match = rail_euro_df[rail_euro_df['Railway station code NTS of origin'] == ori_location]
        if not ori_match.empty:
            ori_code = ori_match['Origin code'].values[0]
            ori_area = ori_match['Origin area'].values[0]
            ori_label = ori_match['Label city of origin '].values[0]

        des_match = rail_euro_df[rail_euro_df['Railway station code NTS of destination'] == des_location]
        if not des_match.empty:
            des_code = des_match['Destination code'].values[0]
            des_country = des_match['Country code of destination'].values[0]
            des_label = des_match['Label city of destination '].values[0]

        return ori_code, ori_label, ori_area, des_code, des_label, des_country

    # Appliquer les requêtes pour les lignes en fonction de la zone
    for idx, row in df.iterrows():
        if row['Zone'] == 'METRO':
            ori_code, ori_label, ori_area, des_code, des_label, des_country = query_metropolis(row['ori_location_rest'], row['des_location_rest'])
        elif row['Zone'] == 'EUROPE':
            ori_code, ori_label, ori_area, des_code, des_label, des_country = query_europe(row['ori_location_rest'], row['des_location_rest'])

        df.at[idx, 'ori_code'] = ori_code
        df.at[idx, 'ori_area'] = ori_area
        df.at[idx, 'des_code'] = des_code
        df.at[idx, 'des_country'] = des_country
        df.at[idx, 'ori_label'] = ori_label
        df.at[idx, 'des_label'] = des_label

    # Sauvegarder dans deux fichiers CSV
    df.to_csv(f"csv/res/train/raw_items_train_{id_soc}.csv", index=False)

def merge_extract_train(id_soc):
    df0 = pd.read_csv(f"csv/base/bills_{id_soc}.csv")
    df0 = df0[df0["type"] == "train"]
    df1 = pd.read_csv(f"csv/res/train/raw_items_train_{id_soc}.csv")
    df = pd.merge(df0, df1, on='ITEM_ID', how='inner')

    # Convert CREATED_AT to YYYY-MM and YYYY formats
    df['ISSUED_MONTH'] = pd.to_datetime(df['CREATED_AT']).dt.strftime('%Y%m')
    df['ODLIST'] = pd.to_datetime(df['CREATED_AT']).dt.strftime('%Y')
    df.to_csv(f"csv/res/train/merge_train_{id_soc}.csv")

def clean_train(id_soc,iata):
    df = pd.read_csv(f'csv/res/train/merge_train_{id_soc}.csv')
    df['O&D'] = df['ori_code'].fillna('') + " " + df['des_code'].fillna('')
    grouped_columns = ['ISSUED_MONTH', 'ODLIST', 'O&D', 'ori_code', 'ori_label',
                       'des_code', 'des_label', 'des_country', 'Zone',"ori_area"]

    grouped_result = df.groupby(grouped_columns).agg(
        TOTAL_BILLED=('TOTAL_BILLED', 'sum'),
        NB_LEGS=('NB_LEGS', 'sum')
    ).reset_index()
    grouped_result = grouped_result.sort_values(by=['O&D', 'ISSUED_MONTH'], ascending=[True, False])

    # Connexion en base pour recup OIN et NAME
    result = col_soc.find_one({"_id": ObjectId(f"{id_soc}")},
                              {"name": 1, "settings.flight.bluebizz": 1, "_id": 0})

    # Extraire les variables "name" et "settings.flight.bluebizz"
    name = result.get("name", None)
    bluebizz = result.get("settings", {}).get("flight", {}).get("bluebizz", None)
    bluebizz_value = bluebizz if bluebizz else "Bluebizz"

    # Remplir les colonnes du DataFrame
    grouped_result['REFERENCE AF'] = bluebizz_value
    grouped_result['RAISON SOCIALE'] = name
    grouped_result['IATA EMETTEUR'] = iata

    #rename col
    grouped_result = grouped_result.rename(columns={
        "ISSUED_MONTH": "DATE D'EMISSION",
        'ori_area': 'ZONE ORIGINE',
        'ori_code': 'CODE ORIGINE',
        'des_code': 'CODE DESTINATION',
        'ori_label': 'LIBELLE ORIGINE',
        'des_label': 'LIBELLE DESTINATION',
        'des_country': 'PAYS DESTINATION',
        'TOTAL_BILLED': 'CA',
        'NB_LEGS': 'NB O&D',
        'des_country': 'PAYS DESTINATION',
    })
    columns_order = [
        "O&D","Zone",'RAISON SOCIALE','REFERENCE AF',"DATE D'EMISSION",'CODE ORIGINE','LIBELLE ORIGINE', 'CODE DESTINATION',"PAYS DESTINATION",'LIBELLE DESTINATION',"IATA EMETTEUR",'ZONE ORIGINE','CA','NB O&D'
    ]

    grouped_result = grouped_result[columns_order]
    grouped_result = grouped_result.sort_values(by=["DATE D'EMISSION","O&D"], ascending=[False, True])

    metro_df = grouped_result[grouped_result['Zone'] == 'METRO']
    europe_df = grouped_result[grouped_result['Zone'] == 'EUROPE']

    # Vérifier les correspondances pour la zone 'METRO'
    metro_df = metro_df[metro_df['Zone'] == 'METRO']
    metro_df['O&D_match'] = metro_df['O&D'].isin(rail_metro_df['O&D restitué'])
    # Filtrer les lignes où 'O&D' a une correspondance
    filtered_metro_df = metro_df[metro_df['O&D_match']].drop(columns=['O&D_match'])

    # Vérifier les correspondances pour la zone 'EUROPE'
    europe_df = europe_df[europe_df['Zone'] == 'EUROPE']
    europe_df['O&D_match'] = europe_df['O&D'].isin(rail_euro_df['O&D restitué'])
    # Filtrer les lignes où 'O&D' a une correspondance
    filtered_europe_df = europe_df[europe_df['O&D_match']].drop(columns=['O&D_match'])


    # Saving the two DataFrames to CSV files
    filtered_metro_df.to_csv(f'csv/res/train/grouped_train_metro_{id_soc}.csv')
    filtered_europe_df.to_csv(f'csv/res/train/grouped_train_euro_{id_soc}.csv')

################ 🤖 TRAITEMENT 🤖 #####################

def calculate_rr_flight_old(id_soc) :

    # Charger les données
    df = pd.read_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv")

    # Convertir 'DATE D\'EMISSION' en datetime et extraire l'année et le mois
    df['DATE D\'EMISSION'] = pd.to_datetime(df['DATE D\'EMISSION'], format='%Y%m')
    df['YEAR'] = df['DATE D\'EMISSION'].dt.year
    df['MONTH'] = df['DATE D\'EMISSION'].dt.month


    # Grouper par O&D, DATE D'EMISSION
    df_grouped = df.groupby(['O&D', 'DATE D\'EMISSION']).agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()


    # Filtrer les données pour "Online" uniquement
    df_online = df[df['TYPE DE VENTE'] == 'Online']

    df_online_grouped = df_online.groupby(['O&D', 'DATE D\'EMISSION']).agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    # Extraire l'année et le mois après le groupement
    df_grouped['YEAR'] = df_grouped['DATE D\'EMISSION'].dt.year
    df_grouped['MONTH'] = df_grouped['DATE D\'EMISSION'].dt.month
    df_online_grouped['YEAR'] = df_online_grouped['DATE D\'EMISSION'].dt.year
    df_online_grouped['MONTH'] = df_online_grouped['DATE D\'EMISSION'].dt.month

    # Définir le mois et l'année précédents
    current_year = pd.Timestamp.now().year
    current_month = pd.Timestamp.now().month

    # Si nous sommes en janvier, le mois précédent est décembre de l'année dernière
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year

    # Filtrer les données du mois N et N-1
    df_n = df_grouped[(df_grouped['YEAR'] == current_year) & (df_grouped['MONTH'] == prev_month)].drop_duplicates(subset=['O&D']).set_index('O&D')
    df_n.to_csv("test.csv")

    # Prendre les données du mois précédent de l'année précédente
    df_n_1 = df_grouped[(df_grouped['YEAR'] == current_year - 1) & (df_grouped['MONTH'] == prev_month)].drop_duplicates(
        subset=['O&D']).set_index('O&D')

    df_online_n = df_online_grouped[(df_online_grouped['YEAR'] == current_year) & (df_online_grouped['MONTH'] == prev_month)].drop_duplicates(subset=['O&D']).set_index('O&D')

    # Calcul des RR pour chaque colonne
    columns_to_calculate = ['CA TOTAL INDUSTRIE', 'CA GROUPE AF KL', 'NB O&D GROUPE AF KL', 'NB O&D INDUSTRIE']
    df_rr_combined = pd.DataFrame(index=df_n.index)
    for column in columns_to_calculate:
        df_rr = df_n[[column]].join(df_n_1[[column]], lsuffix='_N', rsuffix='_N_1')
        df_rr[f'RR_{column}'] = (((df_rr[f'{column}_N'] / df_rr[f'{column}_N_1']) - 1)*100).fillna(0).round(2)
        # Join N-1 and N values to df_rr_combined
        df_rr_combined = df_rr_combined.join(df_rr[[f'{column}_N', f'{column}_N_1', f'RR_{column}']])

    # Calcul des ratios TP CA et TP OD
    df_rr_combined['TP_CA_AF_KL'] = ((df_n['CA GROUPE AF KL'] / df_n['CA TOTAL INDUSTRIE'])*100).round(2)
    df_rr_combined['TP_OD_AF_KL'] = ((df_n['NB O&D GROUPE AF KL'] / df_n['NB O&D INDUSTRIE'])*100).round(2)
    df_rr_combined['TP_CA_AF_KL_N_1'] = ((df_n_1['CA GROUPE AF KL'] / df_n_1['CA TOTAL INDUSTRIE'])*100).round(2)
    df_rr_combined['TP_OD_AF_KL_N_1'] = ((df_n_1['NB O&D GROUPE AF KL'] / df_n_1['NB O&D INDUSTRIE'])*100).round(2)

    # Calcul des nouvelles colonnes pour "Online"
    df_rr_combined['CA_AF_KL_ONLINE'] = df_online_n['CA GROUPE AF KL'].fillna(0)
    df_rr_combined['OD_AF_KL_ONLINE'] = df_online_n['NB O&D GROUPE AF KL'].fillna(0)
    df_rr_combined['CA_INDUSTRIE_ONLINE'] = df_online_n['CA TOTAL INDUSTRIE'].fillna(0)
    df_rr_combined['OD_INDUSTRIE_ONLINE'] = df_online_n['NB O&D INDUSTRIE'].fillna(0)

    # Calcul des ratios "Online"
    df_rr_combined['TP_CA_AF_KL_ONLINE'] = ((df_online_n['CA GROUPE AF KL'] / df_online_n['CA TOTAL INDUSTRIE'])*100).round(2)
    df_rr_combined['TP_OD_AF_KL_ONLINE'] = ((df_online_n['NB O&D GROUPE AF KL'] / df_online_n['NB O&D INDUSTRIE'])*100).round(2)

    # Calcul des évolutions TP CA et TP OD
    df_rr_combined['EVOL_TP_CA'] = (df_rr_combined['TP_CA_AF_KL'] - df_rr_combined['TP_CA_AF_KL_N_1']).round(2)
    df_rr_combined['EVOL_TP_OD'] = (df_rr_combined['TP_OD_AF_KL'] - df_rr_combined['TP_OD_AF_KL_N_1']).round(2)

    #### CUMULS

    # Calculer les cumuls de janvier jusqu'au mois dernier pour 'CA TOTAL INDUSTRIE', 'CA GROUPE AF KL', etc.
    df_cumul = df_grouped[(df_grouped['YEAR'] == current_year) & (df_grouped['MONTH'] <= prev_month)]
    # Grouper par 'O&D' pour calculer les cumuls, sans écraser les colonnes existantes
    df_cumul_grouped = df_cumul.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    # Renommer les colonnes pour indiquer qu'il s'agit des cumuls
    df_cumul_grouped.rename(columns={
        'CA TOTAL INDUSTRIE': 'cumul_CA_TOTAL_INDUSTRIE',
        'CA GROUPE AF KL': 'cumul_CA_GROUPE_AF_KL',
        'NB O&D GROUPE AF KL': 'cumul_NB_O&D_GROUPE_AF_KL',
        'NB O&D INDUSTRIE': 'cumul_NB_O&D_INDUSTRIE'
    }, inplace=True)




    # Calcul des cumuls TP_CA_AF_KL et TP_OD_AF_KL pour les transactions "Online"
    df_online_cumul = df_online_grouped[(df_online_grouped['YEAR'] == current_year) & (df_online_grouped['MONTH'] < current_month)]
    df_online_cumul_grouped = df_online_cumul.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    df_online_cumul_grouped['cumul_TP_CA_AF_KL_ONLINE'] = ((df_online_cumul_grouped['CA GROUPE AF KL'] / df_online_cumul_grouped['CA TOTAL INDUSTRIE'])*100).round(2)
    df_online_cumul_grouped['cumul_TP_OD_AF_KL_ONLINE'] = ((df_online_cumul_grouped['NB O&D GROUPE AF KL'] / df_online_cumul_grouped['NB O&D INDUSTRIE'])*100).round(2)


    # Vérification de l'année en cours et du mois précédent
    current_year = pd.Timestamp.now().year
    current_month = pd.Timestamp.now().month

    # Calcul du mois précédent
    if current_month == 1:
        prev_month = 12  # Si c'est janvier, on passe à décembre
        prev_year = current_year - 1  # Et on soustrait 1 année
    else:
        prev_month = current_month - 1
        prev_year = current_year  # Pas de changement d'année si on est après janvier

    # Vérification de l'année N-1 pour les cumuls
    cumul_prev_year = current_year - 1  # Année N-1 pour les cumuls de janvier à mois courant N-1

    # Filtrer les données du mois N (mois dernier) et N-1 (même mois, année précédente)
    # df_n = df_grouped[(df_grouped['YEAR'] == current_year) & (df_grouped['MONTH'] == prev_month)]
    # df_n_1 = df_grouped[(df_grouped['YEAR'] == prev_year) & (df_grouped['MONTH'] == prev_month)]

    # Filtrer les données pour les cumuls de janvier à mois courant de l'année N-1
    df_cumul_n_1 = df_grouped[(df_grouped['YEAR'] == cumul_prev_year) & (df_grouped['MONTH'] <= prev_month)]

    # Grouper les données pour calculer les cumuls par O&D pour N-1
    df_cumul_n_1_grouped = df_cumul_n_1.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    # Calculer le cumul TP_CA_AF_KL et TP_OD_AF_KL par O&D pour l'année N-1
    df_cumul_grouped['cumul_TP_CA_AF_KL_N_1'] = df_cumul_grouped['O&D'].map(
        df_cumul_n_1_grouped.set_index('O&D').apply(lambda row: (row['CA GROUPE AF KL'] / row['CA TOTAL INDUSTRIE']).round(2) if row['CA TOTAL INDUSTRIE'] > 0 else 0,
            axis=1))
    df_cumul_grouped['cumul_TP_OD_AF_KL_N_1'] = df_cumul_grouped['O&D'].map(df_cumul_n_1_grouped.set_index('O&D').apply(lambda row: (row['NB O&D GROUPE AF KL'] / row['NB O&D INDUSTRIE']).round(2) if row['NB O&D INDUSTRIE'] > 0 else 0,
            axis=1))

    # Calcul des ratios TP_CA et TP_OD pour les cumuls
    df_cumul_grouped['cumul_TP_CA_AF_KL'] = (df_cumul_grouped['cumul_CA_GROUPE_AF_KL'] / df_cumul_grouped['cumul_CA_TOTAL_INDUSTRIE']).round(2)
    df_cumul_grouped['cumul_TP_OD_AF_KL'] = (df_cumul_grouped['cumul_NB_O&D_GROUPE_AF_KL'] / df_cumul_grouped['cumul_NB_O&D_INDUSTRIE']).round(2)

    # Calcul des cumuls RR pour 'CA TOTAL INDUSTRIE' et 'NB O&D INDUSTRIE'
    df_cumul_grouped['cumul_RR_CA TOTAL INDUSTRIE'] = (((df_cumul_grouped['cumul_TP_CA_AF_KL']/df_cumul_grouped['cumul_TP_CA_AF_KL_N_1'])-1)*100).round(2)
    df_cumul_grouped['cumul_RR_NB O&D INDUSTRIE'] = (df_cumul_grouped['cumul_NB_O&D_INDUSTRIE'].pct_change()).round(2)

    # Calcul des cumuls RR pour 'CA GROUPE AF KL' et 'NB O&D GROUPE AF KL'
    df_cumul_grouped['cumul_RR_CA GROUPE AF KL'] = (df_cumul_grouped['cumul_CA_GROUPE_AF_KL'].pct_change()).round(2)
    df_cumul_grouped['cumul_RR_NB O&D GROUPE AF KL'] = (df_cumul_grouped['cumul_NB_O&D_GROUPE_AF_KL'].pct_change()).round(2)


    # Calcul des cumuls EVOL_TP_CA et EVOL_TP_OD en utilisant les cumuls
    df_cumul_grouped['cumul_EVOL_TP_CA'] = (df_cumul_grouped['cumul_TP_CA_AF_KL'] - df_cumul_grouped['cumul_TP_CA_AF_KL_N_1']).round(2)
    df_cumul_grouped['cumul_EVOL_TP_OD'] = (df_cumul_grouped['cumul_TP_OD_AF_KL'] - df_cumul_grouped['cumul_TP_OD_AF_KL_N_1']).round(2)

    # Joindre les résultats "Online" à df_cumul_grouped sans écraser les colonnes
    df_cumul_grouped = df_cumul_grouped.join(df_online_cumul_grouped.set_index('O&D')[['cumul_TP_CA_AF_KL_ONLINE', 'cumul_TP_OD_AF_KL_ONLINE']], on='O&D')

    # Filtrer les données pour l'année N-1
    df_cumul_n_1 = df_grouped[(df_grouped['YEAR'] == current_year - 1) & (df_grouped['MONTH'] <= prev_month)]

    # Grouper par 'O&D' pour calculer les cumuls pour N-1
    df_cumul_n_1_grouped = df_cumul_n_1.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()


    # Renommer les colonnes pour indiquer qu'il s'agit des cumuls N-1
    df_cumul_n_1_grouped.rename(columns={
        'CA TOTAL INDUSTRIE': 'cumul_CA_TOTAL_INDUSTRIE_N_1',
        'CA GROUPE AF KL': 'cumul_CA_GROUPE_AF_KL_N_1',
        'NB O&D GROUPE AF KL': 'cumul_NB_O&D_GROUPE_AF_KL_N_1',
        'NB O&D INDUSTRIE': 'cumul_NB_O&D_INDUSTRIE_N_1'
    }, inplace=True)

    # Calculer les cumuls pour les données "Online"
    df_online_cumul = df_online_grouped[
        (df_online_grouped['YEAR'] == current_year) & (df_online_grouped['MONTH'] <= prev_month)]
    df_online_cumul_grouped = df_online_cumul.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    # Renommer les colonnes pour indiquer qu'il s'agit des cumuls online
    df_online_cumul_grouped.rename(columns={
        'CA TOTAL INDUSTRIE': 'cumul_CA_INDUSTRIE_ONLINE',
        'CA GROUPE AF KL': 'cumul_CA_AF_KL_ONLINE',
        'NB O&D GROUPE AF KL': 'cumul_OD_AF_KL_ONLINE',
        'NB O&D INDUSTRIE': 'cumul_OD_INDUSTRIE_ONLINE'
    }, inplace=True)

    df_cumul_n_1 = df_grouped[(df_grouped['YEAR'] == current_year - 1) & (df_grouped['MONTH'] <= prev_month)]
    df_cumul_n_1_grouped = df_cumul_n_1.groupby('O&D').agg({
        'CA TOTAL INDUSTRIE': 'sum',
        'CA GROUPE AF KL': 'sum',
        'NB O&D GROUPE AF KL': 'sum',
        'NB O&D INDUSTRIE': 'sum'
    }).reset_index()

    # Renommer les colonnes pour indiquer qu'il s'agit des cumuls pour N-1
    df_cumul_n_1_grouped.rename(columns={
        'CA TOTAL INDUSTRIE': 'cumul_CA_TOTAL_INDUSTRIE_N_1',
        'CA GROUPE AF KL': 'cumul_CA_GROUPE_AF_KL_N_1',
        'NB O&D GROUPE AF KL': 'cumul_NB_O&D_GROUPE_AF_KL_N_1',
        'NB O&D INDUSTRIE': 'cumul_NB_O&D_INDUSTRIE_N_1'
    }, inplace=True)

    # Fusionner les cumuls actuels et ceux de N-1 (écraser les colonnes existantes)
    df_cumul_grouped = df_cumul_grouped.merge(df_cumul_n_1_grouped, on='O&D', how='left')

    # Fusionner les résultats "Online" avec df_cumul_grouped (écraser les colonnes existantes)
    df_cumul_grouped = df_cumul_grouped.merge(df_online_cumul_grouped, on='O&D', how='left')

    # Fusionner les résultats de RR, TP et cumuls avec df_rr_combined (écraser les colonnes existantes)
    df_final = pd.merge(df_rr_combined, df_cumul_grouped, on='O&D', how='outer')

    # Sauvegarder le fichier final avec RR, ratios et cumuls
    df_final.to_csv(f"csv/res/flight/grouped_flight_with_rr_{id_soc}.csv")

def calculate_rr_flight_cumul(id_soc):
        # Charger les données
        df = pd.read_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv")

        # Convertir la colonne "DATE D'EMISSION" en datetime
        df["DATE D'EMISSION"] = pd.to_datetime(df["DATE D'EMISSION"], format='%Y%m')

        # Convertir les colonnes "ODLIST" et "IATA EMETTEUR" en texte pour les exclure des calculs numériques
        df['ODLIST'] = df['ODLIST'].astype(str)
        df['IATA EMETTEUR'] = df['IATA EMETTEUR'].astype(str)

        # Obtenir l'année en cours et le mois précédent
        current_year = datetime.now().year
        last_month = datetime.now().month - 1
        if last_month == 0:
            last_month = 12
            current_year -= 1

        # Définir les périodes de filtre
        start_date_n = f"{current_year}-01-01"
        end_date_n = f"{current_year}-{last_month:02d}-28"
        start_date_n_1 = f"{current_year - 1}-01-01"
        end_date_n_1 = f"{current_year - 1}-{last_month:02d}-28"

        # Filtrer les données pour les périodes souhaitées
        df_n = df[
            (df["DATE D'EMISSION"] >= start_date_n) & (df["DATE D'EMISSION"] <= end_date_n) & (df['ODLIST'] == '2024')]
        df_n_1 = df[(df["DATE D'EMISSION"] >= start_date_n_1) & (df["DATE D'EMISSION"] <= end_date_n_1) & (
                    df['ODLIST'] == '2023')]

        # Filtrer les données pour "Online"
        df_n_online = df_n[df_n["TYPE DE VENTE"] == "Online"]
        df_n_1_online = df_n_1[df_n_1["TYPE DE VENTE"] == "Online"]

        # Group by the specified columns (exclure "ODLIST" et "IATA EMETTEUR")
        grouped_columns = [
            'O&D', 'ZONE ORIGINE', 'PERIMETRE', 'LIBELLE ORIGINE', 'LIBELLE DESTINATION',
            'CODE ORIGINE', 'CODE DESTINATION', 'REFERENCE AF', 'RAISON SOCIALE'
        ]

        # Effectuer les opérations d'agrégation uniquement sur les colonnes numériques
        numeric_columns = df_n.select_dtypes(include=['number']).columns

        grouped_n = df_n.groupby(grouped_columns)[numeric_columns].sum().reset_index()
        grouped_n_1 = df_n_1.groupby(grouped_columns)[numeric_columns].sum().reset_index()

        # Grouper les données "Online"
        grouped_n_online = df_n_online.groupby(grouped_columns)[numeric_columns].sum().reset_index()
        grouped_n_1_online = df_n_1_online.groupby(grouped_columns)[numeric_columns].sum().reset_index()

        # Merge the results for both years to ensure they align properly
        result = pd.merge(grouped_n, grouped_n_1, on=grouped_columns, how='outer', suffixes=('_N', '_N_1')).fillna(0)
        result_online = pd.merge(grouped_n_online, grouped_n_1_online, on=grouped_columns, how='outer',
                                 suffixes=('_N', '_N_1')).fillna(0)

        # Calculs des colonnes cumulées avant de procéder aux autres calculs
        result['cumul_OD_AFKL_N'] = result['NB O&D GROUPE AF KL_N']
        result['cumul_CA_AFKL_N'] = result['CA GROUPE AF KL_N']
        result['cumul_OD_INDUSTRIE_N'] = result['NB O&D INDUSTRIE_N']
        result['cumul_CA_INDUSTRIE_N'] = result['CA TOTAL INDUSTRIE_N']

        result['cumul_OD_AFKL_N_1'] = result['NB O&D GROUPE AF KL_N_1']
        result['cumul_CA_AFKL_N_1'] = result['CA GROUPE AF KL_N_1']
        result['cumul_OD_INDUSTRIE_N_1'] = result['NB O&D INDUSTRIE_N_1']
        result['cumul_CA_INDUSTRIE_N_1'] = result['CA TOTAL INDUSTRIE_N_1']

        # Calcul des colonnes pour "Online"
        result['cumul_OD_AFKL_ONLINE_N'] = result_online['NB O&D GROUPE AF KL_N']
        result['cumul_CA_AFKL_ONLINE_N'] = result_online['CA GROUPE AF KL_N']
        result['cumul_OD_INDUSTRIE_ONLINE_N'] = result_online['NB O&D INDUSTRIE_N']
        result['cumul_CA_INDUSTRIE_ONLINE_N'] = result_online['CA TOTAL INDUSTRIE_N']

        result['cumul_OD_AFKL_ONLINE_N_1'] = result_online['NB O&D GROUPE AF KL_N_1']
        result['cumul_CA_AFKL_ONLINE_N_1'] = result_online['CA GROUPE AF KL_N_1']
        result['cumul_OD_INDUSTRIE_ONLINE_N_1'] = result_online['NB O&D INDUSTRIE_N_1']
        result['cumul_CA_INDUSTRIE_ONLINE_N_1'] = result_online['CA TOTAL INDUSTRIE_N_1']

        # Adding the new calculated columns with the specified formulas for TP and RR
        result['cumul_TP_OD_AFKL_N'] = ((result['cumul_OD_AFKL_N'] / result['cumul_OD_INDUSTRIE_N']) * 100).round(2)
        result['cumul_TP_CA_AFKL_N'] = ((result['cumul_CA_AFKL_N'] / result['cumul_CA_INDUSTRIE_N']) * 100).round(2)
        result['cumul_TP_OD_AFKL_ONLINE_N'] = (
                (result['cumul_OD_AFKL_ONLINE_N'] / result['cumul_OD_INDUSTRIE_ONLINE_N']) * 100).round(2)
        result['cumul_TP_CA_AFKL_ONLINE_N'] = (
                (result['cumul_CA_AFKL_ONLINE_N'] / result['cumul_CA_INDUSTRIE_ONLINE_N']) * 100).round(2)

        result['cumul_TP_OD_AFKL_N_1'] = ((result['cumul_OD_AFKL_N_1'] / result['cumul_OD_INDUSTRIE_N_1']) * 100).round(
            2)
        result['cumul_TP_CA_AFKL_N_1'] = ((result['cumul_CA_AFKL_N_1'] / result['cumul_CA_INDUSTRIE_N_1']) * 100).round(
            2)
        result['cumul_TP_OD_AFKL_ONLINE_N_1'] = (
                (result['cumul_OD_AFKL_ONLINE_N_1'] / result['cumul_OD_INDUSTRIE_ONLINE_N_1']) * 100).round(2)
        result['cumul_TP_CA_AFKL_ONLINE_N_1'] = (
                (result['cumul_CA_AFKL_ONLINE_N_1'] / result['cumul_CA_INDUSTRIE_ONLINE_N_1']) * 100).round(2)

        result['cumul_RR_OD_AFKL_N'] = ((result['cumul_OD_AFKL_N'] / result['cumul_OD_AFKL_N_1'] - 1) * 100).round(2)
        result['cumul_RR_CA_AFKL_N'] = ((result['cumul_CA_AFKL_N'] / result['cumul_CA_AFKL_N_1'] - 1) * 100).round(2)
        result['cumul_RR_OD_INDUSTRIE_N'] = (
                (result['cumul_OD_INDUSTRIE_N'] / result['cumul_OD_INDUSTRIE_N_1'] - 1) * 100).round(2)
        result['cumul_RR_CA_INDUSTRIE_N'] = (
                (result['cumul_CA_INDUSTRIE_N'] / result['cumul_CA_INDUSTRIE_N_1'] - 1) * 100).round(2)

        result['cumul_EVOL_TP_OD'] = (result['cumul_TP_OD_AFKL_N'] - result['cumul_TP_OD_AFKL_N_1']).round(2)
        result['cumul_EVOL_TP_CA'] = (result['cumul_TP_CA_AFKL_N'] - result['cumul_TP_CA_AFKL_N_1']).round(2)

        columns_to_drop = [
            'CA GROUPE AF KL_N', 'CA TOTAL INDUSTRIE_N', 'NB O&D GROUPE AF KL_N', 'NB O&D INDUSTRIE_N',
            'CA GROUPE AF KL_N_1', 'CA TOTAL INDUSTRIE_N_1', 'NB O&D GROUPE AF KL_N_1', 'NB O&D INDUSTRIE_N_1'
        ]

        result = result.drop(columns=columns_to_drop)

        # Save the result to a new CSV file
        output_path = f'csv/res/flight/flight_RR_cumul_{id_soc}.csv'
        result.to_csv(output_path, index=False)

def calculate_rr_flight_current(id_soc):
    # Charger les données
    df = pd.read_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv")

    # Convertir la colonne "DATE D'EMISSION" en datetime
    df["DATE D'EMISSION"] = pd.to_datetime(df["DATE D'EMISSION"], format='%Y%m')

    # Convertir les colonnes "ODLIST" et "IATA EMETTEUR" en texte pour les exclure des calculs numériques
    df['ODLIST'] = df['ODLIST'].astype(str)
    df['IATA EMETTEUR'] = df['IATA EMETTEUR'].astype(str)

    # Obtenir l'année en cours et le mois précédent
    current_year = datetime.now().year
    last_month = datetime.now().month - 1
    if last_month == 0:
        last_month = 12
        current_year -= 1

    # Définir les périodes de filtre
    start_date_last_month = f"{current_year}-{last_month:02d}-01"
    end_date_last_month = f"{current_year}-{last_month:02d}-28"

    start_date_last_month_n_1 = f"{current_year - 1}-{last_month:02d}-01"
    end_date_last_month_n_1 = f"{current_year - 1}-{last_month:02d}-28"

    # Filtrer les données pour le mois dernier (N) et le mois équivalent de l'année précédente (N-1)
    df_n = df[
        (df["DATE D'EMISSION"] >= start_date_last_month) &
        (df["DATE D'EMISSION"] <= end_date_last_month) &
        (df['ODLIST'] == '2024')
        ]
    df_n_1 = df[
        (df["DATE D'EMISSION"] >= start_date_last_month_n_1) &
        (df["DATE D'EMISSION"] <= end_date_last_month_n_1) &
        (df['ODLIST'] == '2023')
        ]

    # Filtrer les données pour "Online"
    df_n_online = df_n[df_n["TYPE DE VENTE"] == "Online"]
    df_n_1_online = df_n_1[df_n_1["TYPE DE VENTE"] == "Online"]

    # Group by the specified columns (exclure "ODLIST" et "IATA EMETTEUR")
    grouped_columns = [
        'O&D', 'ZONE ORIGINE', 'PERIMETRE', 'LIBELLE ORIGINE', 'LIBELLE DESTINATION',
        'CODE ORIGINE', 'CODE DESTINATION', 'REFERENCE AF', 'RAISON SOCIALE'
    ]

    # Effectuer les opérations d'agrégation uniquement sur les colonnes numériques
    numeric_columns = df_n.select_dtypes(include=['number']).columns

    grouped_n = df_n.groupby(grouped_columns)[numeric_columns].sum().reset_index()
    grouped_n_1 = df_n_1.groupby(grouped_columns)[numeric_columns].sum().reset_index()

    # Grouper les données "Online"
    grouped_n_online = df_n_online.groupby(grouped_columns)[numeric_columns].sum().reset_index()
    grouped_n_1_online = df_n_1_online.groupby(grouped_columns)[numeric_columns].sum().reset_index()

    # Merge the results for both years to ensure they align properly
    result = pd.merge(grouped_n, grouped_n_1, on=grouped_columns, how='outer', suffixes=('_N', '_N_1')).fillna(0)
    result_online = pd.merge(grouped_n_online, grouped_n_1_online, on=grouped_columns, how='outer',
                             suffixes=('_N', '_N_1')).fillna(0)

    # Calculs des colonnes currentées avant de procéder aux autres calculs
    result['current_OD_AFKL_N'] = result['NB O&D GROUPE AF KL_N']
    result['current_CA_AFKL_N'] = result['CA GROUPE AF KL_N']
    result['current_OD_INDUSTRIE_N'] = result['NB O&D INDUSTRIE_N']
    result['current_CA_INDUSTRIE_N'] = result['CA TOTAL INDUSTRIE_N']

    result['current_OD_AFKL_N_1'] = result['NB O&D GROUPE AF KL_N_1']
    result['current_CA_AFKL_N_1'] = result['CA GROUPE AF KL_N_1']
    result['current_OD_INDUSTRIE_N_1'] = result['NB O&D INDUSTRIE_N_1']
    result['current_CA_INDUSTRIE_N_1'] = result['CA TOTAL INDUSTRIE_N_1']

    # Calcul des colonnes pour "Online"
    result['current_OD_AFKL_ONLINE_N'] = result_online['NB O&D GROUPE AF KL_N']
    result['current_CA_AFKL_ONLINE_N'] = result_online['CA GROUPE AF KL_N']
    result['current_OD_INDUSTRIE_ONLINE_N'] = result_online['NB O&D INDUSTRIE_N']
    result['current_CA_INDUSTRIE_ONLINE_N'] = result_online['CA TOTAL INDUSTRIE_N']

    result['current_OD_AFKL_ONLINE_N_1'] = result_online['NB O&D GROUPE AF KL_N_1']
    result['current_CA_AFKL_ONLINE_N_1'] = result_online['CA GROUPE AF KL_N_1']
    result['current_OD_INDUSTRIE_ONLINE_N_1'] = result_online['NB O&D INDUSTRIE_N_1']
    result['current_CA_INDUSTRIE_ONLINE_N_1'] = result_online['CA TOTAL INDUSTRIE_N_1']

    # Adding the new calculated columns with the specified formulas for TP and RR
    result['current_TP_OD_AFKL_N'] = ((result['current_OD_AFKL_N'] / result['current_OD_INDUSTRIE_N']) * 100).round(2)
    result['current_TP_CA_AFKL_N'] = ((result['current_CA_AFKL_N'] / result['current_CA_INDUSTRIE_N']) * 100).round(2)
    result['current_TP_OD_AFKL_ONLINE_N'] = (
            (result['current_OD_AFKL_ONLINE_N'] / result['current_OD_INDUSTRIE_ONLINE_N']) * 100).round(2)
    result['current_TP_CA_AFKL_ONLINE_N'] = (
            (result['current_CA_AFKL_ONLINE_N'] / result['current_CA_INDUSTRIE_ONLINE_N']) * 100).round(2)

    result['current_TP_OD_AFKL_N_1'] = (
                (result['current_OD_AFKL_N_1'] / result['current_OD_INDUSTRIE_N_1']) * 100).round(2)
    result['current_TP_CA_AFKL_N_1'] = (
                (result['current_CA_AFKL_N_1'] / result['current_CA_INDUSTRIE_N_1']) * 100).round(2)
    result['current_TP_OD_AFKL_ONLINE_N_1'] = (
            (result['current_OD_AFKL_ONLINE_N_1'] / result['current_OD_INDUSTRIE_ONLINE_N_1']) * 100).round(2)
    result['current_TP_CA_AFKL_ONLINE_N_1'] = (
            (result['current_CA_AFKL_ONLINE_N_1'] / result['current_CA_INDUSTRIE_ONLINE_N_1']) * 100).round(2)

    result['current_RR_OD_AFKL_N'] = ((result['current_OD_AFKL_N'] / result['current_OD_AFKL_N_1'] - 1) * 100).round(2)
    result['current_RR_CA_AFKL_N'] = ((result['current_CA_AFKL_N'] / result['current_CA_AFKL_N_1'] - 1) * 100).round(2)
    result['current_RR_OD_INDUSTRIE_N'] = (
            (result['current_OD_INDUSTRIE_N'] / result['current_OD_INDUSTRIE_N_1'] - 1) * 100).round(2)
    result['current_RR_CA_INDUSTRIE_N'] = (
            (result['current_CA_INDUSTRIE_N'] / result['current_CA_INDUSTRIE_N_1'] - 1) * 100).round(2)

    result['current_EVOL_TP_OD'] = (result['current_TP_OD_AFKL_N'] - result['current_TP_OD_AFKL_N_1']).round(2)
    result['current_EVOL_TP_CA'] = (result['current_TP_CA_AFKL_N'] - result['current_TP_CA_AFKL_N_1']).round(2)


    columns_to_drop = [
        'CA GROUPE AF KL_N', 'CA TOTAL INDUSTRIE_N', 'NB O&D GROUPE AF KL_N', 'NB O&D INDUSTRIE_N',
        'CA GROUPE AF KL_N_1', 'CA TOTAL INDUSTRIE_N_1', 'NB O&D GROUPE AF KL_N_1', 'NB O&D INDUSTRIE_N_1'
    ]

    result = result.drop(columns=columns_to_drop)

    # result["DATE D'EMISSION"] = result["DATE D'EMISSION"].dt.strftime('%Y%m')

    # Save the result to a new CSV file
    output_path = f'csv/res/flight/flight_RR_current_{id_soc}.csv'
    result.to_csv(output_path, index=False)

def merge_current_cumul(id_soc):
    file_cumul = f"csv/res/flight/flight_RR_cumul_{id_soc}.csv"
    file_current = f"csv/res/flight/flight_RR_current_{id_soc}.csv"

    # Lire les fichiers en DataFrames
    df_cumul = pd.read_csv(file_cumul)
    df_current = pd.read_csv(file_current)

    # Perform the merge using 'O&D' and 'ODLIST' as keys
    merged_df = pd.merge(df_cumul, df_current, on=['O&D'], how='left', suffixes=('', '_cumul'))

    # Drop duplicated columns that exist in both dataframes from the cumulative data (after merging)
    merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

    # Enregistrer le DataFrame fusionné dans un nouveau fichier Excel
    output_path = f"csv/res/flight/flight_RR_merged_{id_soc}.csv"
    merged_df.to_csv(output_path, index=False)
def merge_rr(id_soc):
    import pandas as pd

    # Load grouped_flight_rr and filtered_flight datasets
    grouped_flight_rr = pd.read_csv(f'csv/res/flight/flight_RR_merged_{id_soc}.csv')

    # Initialize new columns for the required details
    grouped_flight_rr['HAUL_TYPE'] = None
    grouped_flight_rr['ORIGIN_AREA'] = None
    grouped_flight_rr['ANNEX_C'] = None
    grouped_flight_rr['LABEL_ORIGIN'] = None
    grouped_flight_rr['LABEL_DESTINATION'] = None
    grouped_flight_rr['COUNTRY_OF_DEST'] = None  # New column for 'Label country of destination'

    # Use .loc to match and assign all requested columns where O&D matches CONCATENATED_CITIES
    for idx in grouped_flight_rr.index:
        od_value = grouped_flight_rr.loc[idx, 'O&D']
        match = airports_df.loc[airports_df['O&D restitué'] == od_value]

        if not match.empty:
            grouped_flight_rr.loc[idx, 'HAUL_TYPE'] = match['Haul type'].values[0]
            grouped_flight_rr.loc[idx, 'ORIGIN_AREA'] = match['Origin area'].values[0]
            grouped_flight_rr.loc[idx, 'ANNEX_C'] = match['Annex C/ Out of Annex C 2023'].values[0]
            grouped_flight_rr.loc[idx, 'LABEL_ORIGIN'] = match['Label cities of origin'].values[0]
            grouped_flight_rr.loc[idx, 'LABEL_DESTINATION'] = match['Label cities of destination'].values[0]
            grouped_flight_rr.loc[idx, 'COUNTRY_OF_DEST'] = match['Country code of destination'].values[0]


    # Split 'O&D' into 'ORI' and 'DEST'
    grouped_flight_rr[['ORI', 'DEST']] = grouped_flight_rr['O&D'].str.split(n=1, expand=True)


    grouped_flight_rr.to_csv(f'csv/res/flight/flight_full_{id_soc}.csv', index=False)

def split_final(id_soc):
    import pandas as pd

    # Load the input CSV files
    df = pd.read_csv(f"csv/res/flight/flight_full_{id_soc}.csv")

    # Define the desired column order
    columns_order = [
        'O&D', 'ANNEX_C', 'ORI', 'DEST', 'LABEL_ORIGIN', 'LABEL_DESTINATION', 'COUNTRY_OF_DEST',
        'HAUL_TYPE', 'ORIGIN_AREA',

        'current_CA_INDUSTRIE_N', 'current_RR_CA_INDUSTRIE_N',
        'current_OD_INDUSTRIE_N', 'current_RR_OD_INDUSTRIE_N',
        'current_CA_AFKL_N', 'current_RR_CA_AFKL_N',
        'current_OD_AFKL_N', 'current_RR_OD_AFKL_N',
        'current_TP_CA_AFKL_N', 'current_TP_CA_AFKL_ONLINE_N','current_EVOL_TP_CA',
        'current_TP_OD_AFKL_N', 'current_TP_OD_AFKL_ONLINE_N', 'current_EVOL_TP_OD',

        'cumul_CA_INDUSTRIE_N', 'cumul_RR_CA_INDUSTRIE_N',
        'cumul_OD_INDUSTRIE_N', 'cumul_RR_OD_INDUSTRIE_N',
        'cumul_CA_AFKL_N', 'cumul_RR_CA_AFKL_N',
        'cumul_OD_AFKL_N', 'cumul_RR_OD_AFKL_N',
        'cumul_TP_CA_AFKL_N', 'cumul_TP_CA_AFKL_ONLINE_N', 'cumul_EVOL_TP_CA',
        'cumul_TP_OD_AFKL_N', 'cumul_TP_OD_AFKL_ONLINE_N', 'cumul_EVOL_TP_OD',
    ]

    # Split the DataFrame based on the 'ANNEX_C' column
    df_split = dict(tuple(df.groupby('ANNEX_C')))

    # Iterate through the split DataFrames and save each part as a CSV file locally, reordering columns
    for key, df_part in df_split.items():
        # Reorder columns based on the predefined order
        df_part = df_part.reindex(columns=columns_order)

        # Create the filename based on the 'ANNEX_C' value
        filename = f"csv/OK/{key}.csv"

        # Save the DataFrame part with totals to a CSV file locally
        df_part.to_csv(filename, index=False)

    print("Files saved successfully with reordered columns for each split file!")

def aggreg_flight(id_soc,name_orga):
    df = pd.read_csv(f"csv/res/flight/grouped_flight_{id_soc}.csv")

    # Reorder the columns based on the order you provided
    columns_order = [
        'O&D', 'RAISON SOCIALE', 'REFERENCE AF', 'DATE D\'EMISSION',
        'CODE ORIGINE', 'LIBELLE ORIGINE', 'CODE DESTINATION', 'LIBELLE DESTINATION',
        'PERIMETRE', 'IATA EMETTEUR', 'TYPE DE VENTE', 'ZONE ORIGINE',
        'CA TOTAL INDUSTRIE', 'NB O&D INDUSTRIE', 'CA GROUPE AF KL', 'NB O&D GROUPE AF KL'
    ]
    # Reorder the DataFrame
    df = df[columns_order]
    df = df.sort_values(by=['O&D', 'DATE D\'EMISSION'], ascending = [True, False])
    df = df.drop(columns=['O&D'])
    # Save the result to a new CSV file
    df.to_excel(f"csv/OK/Aggregated-IATA_{name_orga}_air.xlsx",index = False)

def aggreg_train(id_soc,name_orga):
    df_euro = pd.read_csv(f"csv/res/train/grouped_train_euro_{id_soc}.csv")
    df_metro = pd.read_csv(f"csv/res/train/grouped_train_metro_{id_soc}.csv")


    # Define the desired column order
    columns_order = [
        'O&D', 'Zone', 'RAISON SOCIALE', 'REFERENCE AF', 'DATE D\'EMISSION',
        'CODE ORIGINE', 'LIBELLE ORIGINE', 'CODE DESTINATION', 'LIBELLE DESTINATION',
        'PAYS DESTINATION', 'IATA EMETTEUR', 'ZONE ORIGINE', 'CA', 'NB O&D'
    ]

    df_euro = df_euro.reindex(columns=columns_order)
    df_metro = df_metro.reindex(columns=columns_order)

    df_metro = df_metro.drop(columns=['O&D',"Zone"])
    df_euro = df_euro.drop(columns=['O&D',"Zone"])

    # Save the results to new CSV files
    df_euro.to_excel(f"csv/OK/Aggregated-IATA_{name_orga}_rail_euro.xlsx", index=False)
    df_metro.to_excel(f"csv/OK/Aggregated-IATA_{name_orga}_rail_metro.xlsx", index=False)

    print("Reorganization and saving completed for both df_euro and df_metro.")

def calculate_rr_train(id_soc, df_name):
    # Load the CSV file into a DataFrame
    df = pd.read_csv(f'csv/res/train/grouped_train_{df_name}_{id_soc}.csv')

    # Check if the DataFrame is empty
    if df.empty:
        print(f"The {df_name} dataset is empty. No calculations can be performed.")
        return

    # Convert 'DATE D'EMISSION' to datetime using the format '%Y%m'
    df["DATE D'EMISSION"] = pd.to_datetime(df["DATE D'EMISSION"], format='%Y%m')

    # Define current year and month
    current_year = pd.Timestamp.now().year
    current_month = pd.Timestamp.now().month

    # Determine previous month
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year

    # Group by the specified columns
    group_columns = ["O&D", "CODE ORIGINE", "CODE DESTINATION", "LIBELLE ORIGINE", "LIBELLE DESTINATION"]

    # Column name for O&D
    od_column = 'NB O&D'

    # For the previous month of this year (N)
    df_prev_month_n = df[(df["DATE D'EMISSION"].dt.year == current_year) &
                         (df["DATE D'EMISSION"].dt.month == prev_month)].groupby(group_columns).agg(
        CA_last_month_N=("CA", "sum"),
        OD_last_month_N=(od_column, "sum")
    ).reset_index()

    # For the previous month of last year (N-1)
    df_prev_month_n_1 = df[(df["DATE D'EMISSION"].dt.year == current_year - 1) &
                           (df["DATE D'EMISSION"].dt.month == prev_month)].groupby(group_columns).agg(
        CA_last_month_N_1=("CA", "sum"),
        OD_last_month_N_1=(od_column, "sum")
    ).reset_index()

    # For January to the previous month of this year (N)
    df_jan_to_prev_month_n = df[(df["DATE D'EMISSION"].dt.year == current_year) &
                                (df["DATE D'EMISSION"].dt.month <= prev_month)].groupby(group_columns).agg(
        CA_jan_to_last_month_N=("CA", "sum"),
        OD_jan_to_last_month_N=(od_column, "sum")
    ).reset_index()

    # For January to the previous month of last year (N-1)
    df_jan_to_prev_month_n_1 = df[(df["DATE D'EMISSION"].dt.year == current_year - 1) &
                                  (df["DATE D'EMISSION"].dt.month <= prev_month)].groupby(group_columns).agg(
        CA_jan_to_last_month_N_1=("CA", "sum"),
        OD_jan_to_last_month_N_1=(od_column, "sum")
    ).reset_index()

    # Merge all results into a single dataframe
    df_aggregated = pd.merge(df_prev_month_n, df_prev_month_n_1, on=group_columns, how='outer')
    df_aggregated = pd.merge(df_aggregated, df_jan_to_prev_month_n, on=group_columns, how='outer')
    df_aggregated = pd.merge(df_aggregated, df_jan_to_prev_month_n_1, on=group_columns, how='outer')

    # Round non-RR columns to integers
    cols_to_round = [
        'CA_last_month_N', 'OD_last_month_N',
        'CA_last_month_N_1', 'OD_last_month_N_1',
        'CA_jan_to_last_month_N', 'OD_jan_to_last_month_N',
        'CA_jan_to_last_month_N_1', 'OD_jan_to_last_month_N_1'
    ]
    df_aggregated[cols_to_round] = df_aggregated[cols_to_round].round(0)

    # Calculate the Rate of Change (RR) for CA and O&D
    df_aggregated['RR_CA'] = (((df_aggregated['CA_last_month_N'] / df_aggregated['CA_last_month_N_1']) - 1)*100).round(2)
    df_aggregated['RR_OD'] = (((df_aggregated['OD_last_month_N'] / df_aggregated['OD_last_month_N_1']) - 1)*100).round(2)

    # Calculate the Cumulative Rate of Change (RR CUMUL) for CA and O&D
    df_aggregated['cumul_RR_CA'] = (((df_aggregated['CA_jan_to_last_month_N'] / df_aggregated['CA_jan_to_last_month_N_1']) - 1)*100).round(2)
    df_aggregated['cumul_RR_OD'] = (((df_aggregated['OD_jan_to_last_month_N'] / df_aggregated['OD_jan_to_last_month_N_1']) - 1)*100).round(2)

    # Calculate the total values for each metric
    total_values = df_aggregated[cols_to_round].sum()

    # Calculate the Rate of Change (RR) using total values
    rr_od_last_month = ((total_values['OD_last_month_N'] / total_values['OD_last_month_N_1'] - 1)*100).round(2)
    rr_ca_last_month = ((total_values['CA_last_month_N'] / total_values['CA_last_month_N_1'] - 1)*100).round(2)
    rr_ca_jan_to_last_month = ((total_values['CA_jan_to_last_month_N'] / total_values['CA_jan_to_last_month_N_1'] - 1)*100).round(2)
    rr_od_jan_to_last_month = ((total_values['OD_jan_to_last_month_N'] / total_values['OD_jan_to_last_month_N_1'] - 1)*100).round(2)

    # Create a DataFrame for the total row
    total_row = pd.DataFrame([{
        'O&D': 'Total',
        'CODE ORIGINE': '',
        'CODE DESTINATION': '',
        'LIBELLE ORIGINE': '',
        'LIBELLE DESTINATION': '',
        'CA_last_month_N': total_values['CA_last_month_N'],
        'OD_last_month_N': total_values['OD_last_month_N'],
        'CA_last_month_N_1': total_values['CA_last_month_N_1'],
        'OD_last_month_N_1': total_values['OD_last_month_N_1'],
        'CA_jan_to_last_month_N': total_values['CA_jan_to_last_month_N'],
        'OD_jan_to_last_month_N': total_values['OD_jan_to_last_month_N'],
        'CA_jan_to_last_month_N_1': total_values['CA_jan_to_last_month_N_1'],
        'OD_jan_to_last_month_N_1': total_values['OD_jan_to_last_month_N_1'],
        'RR_CA': rr_ca_last_month,
        'RR_OD': rr_od_last_month,
        'cumul_RR_CA': rr_ca_jan_to_last_month,
        'cumul_RR_OD': rr_od_jan_to_last_month
    }])

    # Append the total row to the aggregated DataFrame
    df_aggregated = pd.concat([df_aggregated, total_row], ignore_index=True)

    # Save the aggregated DataFrame with the total row to CSV
    df_aggregated.to_csv(f'csv/OK/total_{df_name}.csv', index=False)

def total_old(id_soc):
    # Load the CSV file
    file_path = f'csv/res/flight/flight_full_{id_soc}.csv'
    df = pd.read_csv(file_path)

    # Reorganize columns: move columns ending with "_N_1" to the end
    columns_to_move = [col for col in df.columns if col.endswith('_N_1')]
    columns_remaining = [col for col in df.columns if col not in columns_to_move]
    new_column_order = columns_remaining + columns_to_move

    # Reorganize the dataframe
    df_reorganized = df[new_column_order]

    # Calculate the totals by grouping by the column 'ANNEX_C'
    grouped_totals = df_reorganized.groupby('ANNEX_C').sum(numeric_only=True).reset_index()

    # TP Ratios for cumulative columns
    grouped_totals['cumul_TP_CA_AFKL_N'] = ((grouped_totals['cumul_CA_AFKL_N'] / grouped_totals['cumul_CA_INDUSTRIE_N'])*100).round(2)
    grouped_totals['cumul_TP_CA_AFKL_N_1'] = ((grouped_totals['cumul_CA_AFKL_N_1'] / grouped_totals['cumul_CA_INDUSTRIE_N_1'])*100).round(2)
    grouped_totals['cumul_TP_OD_AFKL_N'] = ((grouped_totals['cumul_OD_AFKL_N'] / grouped_totals['cumul_OD_INDUSTRIE_N'])*100).round(2)
    grouped_totals['cumul_TP_OD_AFKL_N_1'] = ((grouped_totals['cumul_OD_AFKL_N_1'] / grouped_totals['cumul_OD_INDUSTRIE_N_1'])*100).round(2)

    # Evolution calculation for cumulative values
    grouped_totals['cumul_EVOL_TP_CA'] = (grouped_totals['cumul_TP_CA_AFKL_N'] - grouped_totals['cumul_TP_CA_AFKL_N_1']).round(2)
    grouped_totals['cumul_EVOL_TP_OD'] = (grouped_totals['cumul_TP_OD_AFKL_N'] - grouped_totals['cumul_TP_OD_AFKL_N_1']).round(2)

    # Calculate the subtotals RR for each ANNEX_C group
    grouped_totals['cumul_RR_CA_AFKL_N'] = (((grouped_totals['cumul_CA_AFKL_N'] / grouped_totals['cumul_CA_AFKL_N_1']) - 1)*100).round(2)
    grouped_totals['cumul_RR_OD_AFKL_N'] = (((grouped_totals['cumul_OD_AFKL_N'] / grouped_totals['cumul_OD_AFKL_N_1']) - 1)*100).round(2)
    grouped_totals['cumul_RR_CA_INDUSTRIE_N'] = (((grouped_totals['cumul_CA_INDUSTRIE_N'] / grouped_totals['cumul_CA_INDUSTRIE_N_1']) - 1)*100).round(2)
    grouped_totals['cumul_RR_OD_INDUSTRIE_N'] = (((grouped_totals['cumul_OD_INDUSTRIE_N'] / grouped_totals['cumul_OD_INDUSTRIE_N_1']) - 1)*100).round(2)

    # Calculate TP ratios for each ANNEX_C group
    grouped_totals['cumul_TP_CA_AFKL_N'] = ((grouped_totals['cumul_CA_AFKL_N'] / grouped_totals['cumul_CA_INDUSTRIE_N'])*100).round(2)
    grouped_totals['cumul_TP_CA_AFKL_N_1'] = ((grouped_totals['cumul_CA_AFKL_N_1'] / grouped_totals['cumul_CA_INDUSTRIE_N_1'])*100).round(2)
    grouped_totals['cumul_TP_OD_AFKL_N'] = ((grouped_totals['cumul_OD_AFKL_N'] / grouped_totals['cumul_OD_INDUSTRIE_N'])*100).round(2)
    grouped_totals['cumul_TP_OD_AFKL_N_1'] = ((grouped_totals['cumul_OD_AFKL_N_1'] / grouped_totals['cumul_OD_INDUSTRIE_N_1'])*100).round(2)

    # Evolution calculation
    grouped_totals['cumul_EVOL_TP_CA'] = (grouped_totals['cumul_TP_CA_AFKL_N'] - grouped_totals['cumul_TP_CA_AFKL_N_1']).round(2)
    grouped_totals['cumul_EVOL_TP_OD'] = (grouped_totals['cumul_TP_OD_AFKL_N'] - grouped_totals['cumul_TP_OD_AFKL_N_1']).round(2)

    # -----------------------------------
    # Current Calculations (Prefix: 'current_')
    # -----------------------------------

    # TP Ratios for currentative columns
    grouped_totals['current_TP_CA_AFKL_N'] = (
            grouped_totals['current_CA_AFKL_N'] / grouped_totals['current_CA_INDUSTRIE_N']).round(2)
    grouped_totals['current_TP_CA_AFKL_N_1'] = (
            grouped_totals['current_CA_AFKL_N_1'] / grouped_totals['current_CA_INDUSTRIE_N_1']).round(2)
    grouped_totals['current_TP_OD_AFKL_N'] = (
            grouped_totals['current_OD_AFKL_N'] / grouped_totals['current_OD_INDUSTRIE_N']).round(2)
    grouped_totals['current_TP_OD_AFKL_N_1'] = (
            grouped_totals['current_OD_AFKL_N_1'] / grouped_totals['current_OD_INDUSTRIE_N_1']).round(2)

    # Evolution calculation for currentative values
    grouped_totals['current_EVOL_TP_CA'] = (
            grouped_totals['current_TP_CA_AFKL_N'] - grouped_totals['current_TP_CA_AFKL_N_1']).round(2)
    grouped_totals['current_EVOL_TP_OD'] = (
            grouped_totals['current_TP_OD_AFKL_N'] - grouped_totals['current_TP_OD_AFKL_N_1']).round(2)

    # Calculate the subtotals RR for each ANNEX_C group
    grouped_totals['current_RR_CA_AFKL_N'] = (
                ((grouped_totals['current_CA_AFKL_N'] / grouped_totals['current_CA_AFKL_N_1']) - 1) * 100).round(2)
    grouped_totals['current_RR_OD_AFKL_N'] = (
                ((grouped_totals['current_OD_AFKL_N'] / grouped_totals['current_OD_AFKL_N_1']) - 1) * 100).round(2)
    grouped_totals['current_RR_CA_INDUSTRIE_N'] = (
                ((grouped_totals['current_CA_INDUSTRIE_N'] / grouped_totals['current_CA_INDUSTRIE_N_1']) - 1) * 100).round(
        2)
    grouped_totals['current_RR_OD_INDUSTRIE_N'] = (
                ((grouped_totals['current_OD_INDUSTRIE_N'] / grouped_totals['current_OD_INDUSTRIE_N_1']) - 1) * 100).round(
        2)

    # Calculate TP ratios for each ANNEX_C group
    grouped_totals['current_TP_CA_AFKL_N'] = (
                (grouped_totals['current_CA_AFKL_N'] / grouped_totals['current_CA_INDUSTRIE_N']) * 100).round(2)
    grouped_totals['current_TP_CA_AFKL_N_1'] = (
                (grouped_totals['current_CA_AFKL_N_1'] / grouped_totals['current_CA_INDUSTRIE_N_1']) * 100).round(2)
    grouped_totals['current_TP_OD_AFKL_N'] = (
                (grouped_totals['current_OD_AFKL_N'] / grouped_totals['current_OD_INDUSTRIE_N']) * 100).round(2)
    grouped_totals['current_TP_OD_AFKL_N_1'] = (
                (grouped_totals['current_OD_AFKL_N_1'] / grouped_totals['current_OD_INDUSTRIE_N_1']) * 100).round(2)

    # Evolution calculation
    grouped_totals['current_EVOL_TP_CA'] = (
                grouped_totals['current_TP_CA_AFKL_N'] - grouped_totals['current_TP_CA_AFKL_N_1']).round(2)
    grouped_totals['current_EVOL_TP_OD'] = (
                grouped_totals['current_TP_OD_AFKL_N'] - grouped_totals['current_TP_OD_AFKL_N_1']).round(2)

    output_path = 'csv/OK/totals_flight.csv'

    grouped_totals.to_csv(output_path)

import numpy as np

def total(id_soc):
    # Load the uploaded CSV file
    full_file_path = f'csv/res/flight/flight_full_{id_soc}.csv'

    # Read the CSV file into a dataframe
    df_full = pd.read_csv(full_file_path)

    # Split the dataframe by 'PERIMETRE' into separate dataframes
    perimetre_groups = df_full.groupby('PERIMETRE')

    # Define a function to calculate totals for the given columns based on the specified rules
    def calculate_custom_totals(df):
        totals = {}
        # Calculate specific totals according to given rules
        def safe_divide(numerator, denominator):
            return (numerator / denominator * 100).round(2) if denominator != 0 else np.nan

        # Calculs spécifiques pour les colonnes demandées (multiplication par 100 avant round)
        totals['cumul_TP_CA_AFKL_N'] = safe_divide(df['cumul_CA_AFKL_N'].sum(), df['cumul_CA_INDUSTRIE_N'].sum())
        totals['cumul_TP_CA_AFKL_N_1'] = safe_divide(df['cumul_CA_AFKL_N_1'].sum(), df['cumul_CA_INDUSTRIE_N_1'].sum())
        totals['cumul_TP_OD_AFKL_N'] = safe_divide(df['cumul_OD_AFKL_N'].sum(), df['cumul_OD_INDUSTRIE_N'].sum())
        totals['cumul_TP_OD_AFKL_N_1'] = safe_divide(df['cumul_OD_AFKL_N_1'].sum(), df['cumul_OD_INDUSTRIE_N_1'].sum())
        totals['cumul_TP_CA_AFKL_ONLINE_N'] = safe_divide(df['cumul_CA_AFKL_ONLINE_N'].sum(), df['cumul_CA_INDUSTRIE_ONLINE_N'].sum())
        totals['cumul_TP_CA_AFKL_ONLINE_N_1'] = safe_divide(df['cumul_CA_AFKL_ONLINE_N_1'].sum(), df['cumul_CA_INDUSTRIE_ONLINE_N_1'].sum())
        totals['cumul_TP_OD_AFKL_ONLINE_N'] = safe_divide(df['cumul_OD_AFKL_ONLINE_N'].sum(), df['cumul_OD_INDUSTRIE_ONLINE_N'].sum())
        totals['cumul_TP_OD_AFKL_ONLINE_N_1'] = safe_divide(df['cumul_OD_AFKL_ONLINE_N_1'].sum(), df['cumul_OD_INDUSTRIE_ONLINE_N_1'].sum())

        totals['cumul_EVOL_TP_CA'] = (totals['cumul_TP_CA_AFKL_N'] - totals['cumul_TP_CA_AFKL_N_1']).round(2) if not np.isnan(totals['cumul_TP_CA_AFKL_N_1']) else np.nan
        totals['cumul_EVOL_TP_OD'] = (totals['cumul_TP_OD_AFKL_N'] - totals['cumul_TP_OD_AFKL_N_1']).round(2) if not np.isnan(totals['cumul_TP_OD_AFKL_N_1']) else np.nan

        totals['cumul_RR_CA_AFKL_N'] = (safe_divide(df['cumul_CA_AFKL_N'].sum(), df['cumul_CA_AFKL_N_1'].sum()) - 100) if df['cumul_CA_AFKL_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_OD_AFKL_N'] = (safe_divide(df['cumul_OD_AFKL_N'].sum(), df['cumul_OD_AFKL_N_1'].sum()) - 100) if df['cumul_OD_AFKL_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_CA_INDUSTRIE_N'] = (safe_divide(df['cumul_CA_INDUSTRIE_N'].sum(), df['cumul_CA_INDUSTRIE_N_1'].sum()) - 100) if df['cumul_CA_INDUSTRIE_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_OD_INDUSTRIE_N'] = (safe_divide(df['cumul_OD_INDUSTRIE_N'].sum(), df['cumul_OD_INDUSTRIE_N_1'].sum()) - 100) if df['cumul_OD_INDUSTRIE_N_1'].sum() != 0 else np.nan

        totals['current_TP_CA_AFKL_N'] = safe_divide(df['current_CA_AFKL_N'].sum(), df['current_CA_INDUSTRIE_N'].sum())
        totals['current_TP_CA_AFKL_N_1'] = safe_divide(df['current_CA_AFKL_N_1'].sum(), df['current_CA_INDUSTRIE_N_1'].sum())
        totals['current_TP_OD_AFKL_N'] = safe_divide(df['current_OD_AFKL_N'].sum(), df['current_OD_INDUSTRIE_N'].sum())
        totals['current_TP_OD_AFKL_N_1'] = safe_divide(df['current_OD_AFKL_N_1'].sum(), df['current_OD_INDUSTRIE_N_1'].sum())
        totals['current_TP_CA_AFKL_ONLINE_N'] = safe_divide(df['current_CA_AFKL_ONLINE_N'].sum(),df['current_CA_INDUSTRIE_ONLINE_N'].sum())
        totals['current_TP_CA_AFKL_ONLINE_N_1'] = safe_divide(df['current_CA_AFKL_ONLINE_N_1'].sum(),df['current_CA_INDUSTRIE_ONLINE_N_1'].sum())
        totals['current_TP_OD_AFKL_ONLINE_N'] = safe_divide(df['current_OD_AFKL_ONLINE_N'].sum(),df['current_OD_INDUSTRIE_ONLINE_N'].sum())
        totals['current_TP_OD_AFKL_ONLINE_N_1'] = safe_divide(df['current_OD_AFKL_ONLINE_N_1'].sum(),df['current_OD_INDUSTRIE_ONLINE_N_1'].sum())

        totals['current_EVOL_TP_CA'] = (totals['current_TP_CA_AFKL_N'] - totals['current_TP_CA_AFKL_N_1']).round(2) if not np.isnan(totals['current_TP_CA_AFKL_N_1']) else np.nan
        totals['current_EVOL_TP_OD'] = (totals['current_TP_OD_AFKL_N'] - totals['current_TP_OD_AFKL_N_1']).round(2) if not np.isnan(totals['current_TP_OD_AFKL_N_1']) else np.nan

        totals['current_RR_CA_AFKL_N'] = (safe_divide(df['current_CA_AFKL_N'].sum(), df['current_CA_AFKL_N_1'].sum()) - 100) if df['current_CA_AFKL_N_1'].sum() != 0 else np.nan
        totals['current_RR_OD_AFKL_N'] = (safe_divide(df['current_OD_AFKL_N'].sum(), df['current_OD_AFKL_N_1'].sum()) - 100) if df['current_OD_AFKL_N_1'].sum() != 0 else np.nan
        totals['current_RR_CA_INDUSTRIE_N'] = (safe_divide(df['current_CA_INDUSTRIE_N'].sum(), df['current_CA_INDUSTRIE_N_1'].sum()) - 100) if df['current_CA_INDUSTRIE_N_1'].sum() != 0 else np.nan
        totals['current_RR_OD_INDUSTRIE_N'] = (safe_divide(df['current_OD_INDUSTRIE_N'].sum(), df['current_OD_INDUSTRIE_N_1'].sum()) - 100) if df['current_OD_INDUSTRIE_N_1'].sum() != 0 else np.nan

        # Calcul de la somme pour les autres colonnes numériques non spécifiées
        numeric_columns = df.select_dtypes(include=['number']).columns
        for col in numeric_columns:
            if col not in totals:
                totals[col] = df[col].sum()

        return totals

    # Create an Excel writer object to save all 'PERIMETRE' specific dataframes with calculated totals
    output_file_paths = []

    for perimetre, df_perimetre in perimetre_groups:
        # Calculate custom totals for the specific columns
        custom_totals = calculate_custom_totals(df_perimetre)
        df_totals = pd.DataFrame([custom_totals])

        # Concatenate the dataframe with custom totals
        df_with_totals = pd.concat([df_perimetre, df_totals], ignore_index=True)

        # Save each dataframe to a separate Excel file
        output_file_path = f'csv/OK/{perimetre}.csv'
        df_with_totals.to_csv(output_file_path, index=False)
        output_file_paths.append(output_file_path)

    # Provide information on the generated files
    return output_file_paths

def total_global(id_soc):
    # Load the uploaded CSV file
    full_file_path = f'csv/res/flight/flight_full_{id_soc}.csv'

    # Read the CSV file into a dataframe
    df_full = pd.read_csv(full_file_path)

    # Define a function to calculate totals for the given columns based on the specified rules
    def calculate_custom_totals(df):
        totals = {}
        # Calculate specific totals according to given rules
        def safe_divide(numerator, denominator):
            return (numerator / denominator * 100).round(2) if denominator != 0 else np.nan

        # Calculs spécifiques pour les colonnes demandées (multiplication par 100 avant round)
        totals['cumul_TP_CA_AFKL_N'] = safe_divide(df['cumul_CA_AFKL_N'].sum(), df['cumul_CA_INDUSTRIE_N'].sum())
        totals['cumul_TP_CA_AFKL_N_1'] = safe_divide(df['cumul_CA_AFKL_N_1'].sum(), df['cumul_CA_INDUSTRIE_N_1'].sum())
        totals['cumul_TP_OD_AFKL_N'] = safe_divide(df['cumul_OD_AFKL_N'].sum(), df['cumul_OD_INDUSTRIE_N'].sum())
        totals['cumul_TP_OD_AFKL_N_1'] = safe_divide(df['cumul_OD_AFKL_N_1'].sum(), df['cumul_OD_INDUSTRIE_N_1'].sum())
        totals['cumul_TP_CA_AFKL_ONLINE_N'] = safe_divide(df['cumul_CA_AFKL_ONLINE_N'].sum(), df['cumul_CA_INDUSTRIE_ONLINE_N'].sum())
        totals['cumul_TP_CA_AFKL_ONLINE_N_1'] = safe_divide(df['cumul_CA_AFKL_ONLINE_N_1'].sum(), df['cumul_CA_INDUSTRIE_ONLINE_N_1'].sum())
        totals['cumul_TP_OD_AFKL_ONLINE_N'] = safe_divide(df['cumul_OD_AFKL_ONLINE_N'].sum(), df['cumul_OD_INDUSTRIE_ONLINE_N'].sum())
        totals['cumul_TP_OD_AFKL_ONLINE_N_1'] = safe_divide(df['cumul_OD_AFKL_ONLINE_N_1'].sum(), df['cumul_OD_INDUSTRIE_ONLINE_N_1'].sum())

        totals['cumul_EVOL_TP_CA'] = (totals['cumul_TP_CA_AFKL_N'] - totals['cumul_TP_CA_AFKL_N_1']).round(2) if not np.isnan(totals['cumul_TP_CA_AFKL_N_1']) else np.nan
        totals['cumul_EVOL_TP_OD'] = (totals['cumul_TP_OD_AFKL_N'] - totals['cumul_TP_OD_AFKL_N_1']).round(2) if not np.isnan(totals['cumul_TP_OD_AFKL_N_1']) else np.nan

        totals['cumul_RR_CA_AFKL_N'] = (safe_divide(df['cumul_CA_AFKL_N'].sum(), df['cumul_CA_AFKL_N_1'].sum()) - 100) if df['cumul_CA_AFKL_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_OD_AFKL_N'] = (safe_divide(df['cumul_OD_AFKL_N'].sum(), df['cumul_OD_AFKL_N_1'].sum()) - 100) if df['cumul_OD_AFKL_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_CA_INDUSTRIE_N'] = (safe_divide(df['cumul_CA_INDUSTRIE_N'].sum(), df['cumul_CA_INDUSTRIE_N_1'].sum()) - 100) if df['cumul_CA_INDUSTRIE_N_1'].sum() != 0 else np.nan
        totals['cumul_RR_OD_INDUSTRIE_N'] = (safe_divide(df['cumul_OD_INDUSTRIE_N'].sum(), df['cumul_OD_INDUSTRIE_N_1'].sum()) - 100) if df['cumul_OD_INDUSTRIE_N_1'].sum() != 0 else np.nan

        totals['current_TP_CA_AFKL_N'] = safe_divide(df['current_CA_AFKL_N'].sum(), df['current_CA_INDUSTRIE_N'].sum())
        totals['current_TP_CA_AFKL_N_1'] = safe_divide(df['current_CA_AFKL_N_1'].sum(), df['current_CA_INDUSTRIE_N_1'].sum())
        totals['current_TP_OD_AFKL_N'] = safe_divide(df['current_OD_AFKL_N'].sum(), df['current_OD_INDUSTRIE_N'].sum())
        totals['current_TP_OD_AFKL_N_1'] = safe_divide(df['current_OD_AFKL_N_1'].sum(), df['current_OD_INDUSTRIE_N_1'].sum())
        totals['current_TP_CA_AFKL_ONLINE_N'] = safe_divide(df['current_CA_AFKL_ONLINE_N'].sum(), df['current_CA_INDUSTRIE_ONLINE_N'].sum())
        totals['current_TP_CA_AFKL_ONLINE_N_1'] = safe_divide(df['current_CA_AFKL_ONLINE_N_1'].sum(), df['current_CA_INDUSTRIE_ONLINE_N_1'].sum())
        totals['current_TP_OD_AFKL_ONLINE_N'] = safe_divide(df['current_OD_AFKL_ONLINE_N'].sum(), df['current_OD_INDUSTRIE_ONLINE_N'].sum())
        totals['current_TP_OD_AFKL_ONLINE_N_1'] = safe_divide(df['current_OD_AFKL_ONLINE_N_1'].sum(), df['current_OD_INDUSTRIE_ONLINE_N_1'].sum())

        totals['current_EVOL_TP_CA'] = (totals['current_TP_CA_AFKL_N'] - totals['current_TP_CA_AFKL_N_1']).round(2) if not np.isnan(totals['current_TP_CA_AFKL_N_1']) else np.nan
        totals['current_EVOL_TP_OD'] = (totals['current_TP_OD_AFKL_N'] - totals['current_TP_OD_AFKL_N_1']).round(2) if not np.isnan(totals['current_TP_OD_AFKL_N_1']) else np.nan

        totals['current_RR_CA_AFKL_N'] = (safe_divide(df['current_CA_AFKL_N'].sum(), df['current_CA_AFKL_N_1'].sum()) - 100) if df['current_CA_AFKL_N_1'].sum() != 0 else np.nan
        totals['current_RR_OD_AFKL_N'] = (safe_divide(df['current_OD_AFKL_N'].sum(), df['current_OD_AFKL_N_1'].sum()) - 100) if df['current_OD_AFKL_N_1'].sum() != 0 else np.nan
        totals['current_RR_CA_INDUSTRIE_N'] = (safe_divide(df['current_CA_INDUSTRIE_N'].sum(), df['current_CA_INDUSTRIE_N_1'].sum()) - 100) if df['current_CA_INDUSTRIE_N_1'].sum() != 0 else np.nan
        totals['current_RR_OD_INDUSTRIE_N'] = (safe_divide(df['current_OD_INDUSTRIE_N'].sum(), df['current_OD_INDUSTRIE_N_1'].sum()) - 100) if df['current_OD_INDUSTRIE_N_1'].sum() != 0 else np.nan

        # Calcul de la somme pour les autres colonnes numériques non spécifiées
        numeric_columns = df.select_dtypes(include=['number']).columns
        for col in numeric_columns:
            if col not in totals:
                totals[col] = df[col].sum()

        return totals

    # Calculate custom totals for the entire dataframe
    custom_totals = calculate_custom_totals(df_full)
    df_totals = pd.DataFrame([custom_totals])

    # Concatenate the dataframe with custom totals
    df_with_totals = pd.concat([df_full, df_totals], ignore_index=True)

    # Save the dataframe to an Excel file
    output_file_path = f'csv/OK/total_flight.csv'
    df_with_totals.to_csv(output_file_path, index=False)

def create_excel_flight():
    # Load the CSV files into DataFrames
    iac_df = pd.read_csv('csv/OK/IAC.csv')
    ihac_df = pd.read_csv('csv/OK/IHAC.csv')
    mac_df = pd.read_csv('csv/OK/MAC.csv')
    mhac_df = pd.read_csv('csv/OK/MHAC.csv')

    # Drop the "O&D" and "ANNEX_C" columns if they exist in the DataFrames
    columns_to_exclude = ["O&D", "ANNEX_C"]

    for df in [iac_df, ihac_df, mac_df, mhac_df]:
        for column in columns_to_exclude:
            if column in df.columns:
                df.drop(columns=[column], inplace=True)

    # List of columns to keep and reorder
    columns_to_keep = [
        'ORI', 'DEST', 'LABEL_ORIGIN', 'LABEL_DESTINATION', 'COUNTRY_OF_DEST', 'HAUL_TYPE', 'ORIGIN_AREA',
        'current_CA_INDUSTRIE_N', 'current_RR_CA_INDUSTRIE_N', 'current_OD_INDUSTRIE_N', 'current_RR_OD_INDUSTRIE_N',
        'current_CA_AFKL_N', 'current_RR_CA_AFKL_N', 'current_OD_AFKL_N', 'current_RR_OD_AFKL_N',
        'current_TP_CA_AFKL_N', 'current_TP_CA_AFKL_ONLINE_N', 'current_EVOL_TP_CA',
        'current_TP_OD_AFKL_N', 'current_TP_OD_AFKL_ONLINE_N', 'current_EVOL_TP_OD',
        'cumul_CA_INDUSTRIE_N', 'cumul_RR_CA_INDUSTRIE_N', 'cumul_OD_INDUSTRIE_N', 'cumul_RR_OD_INDUSTRIE_N',
        'cumul_CA_AFKL_N', 'cumul_RR_CA_AFKL_N', 'cumul_OD_AFKL_N', 'cumul_RR_OD_AFKL_N',
        'cumul_TP_CA_AFKL_N', 'cumul_TP_CA_AFKL_ONLINE_N', 'cumul_EVOL_TP_CA',
        'cumul_TP_OD_AFKL_N', 'cumul_TP_OD_AFKL_ONLINE_N', 'cumul_EVOL_TP_OD'
    ]

    # Filter and reorder columns for each DataFrame
    filtered_dfs = []
    for df in [iac_df, ihac_df, mac_df, mhac_df]:
        df_filtered = df[[col for col in columns_to_keep if col in df.columns]]
        filtered_dfs.append(df_filtered)

        # Load the existing Excel file
        excel_path = "const/template.xlsx"
        workbook = load_workbook(excel_path)

        # Select the worksheet "Etat 1.1"
        sheet_name = "Etat 1.1"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print(f"Sheet '{sheet_name}' does not exist in the workbook.")
            return

        # Unmerge all cells in the sheet to handle 'MergedCell' issues
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))

        # Define headers to be used for each table
        headers = [
            "ORI", "DEST", "DESTINATION", "ORIGINE", "CODE_PAYS", "TYPE_COURRIER", "ZONE_ORIGINE",
            "CA", "R/R", "O&D", "R/R", "CA", "R/R", "O&D", "R/R", "T/P CA", "T/P CA ONLINE",
            "EVOL T/P (CA)", "T/P O&D", "T/P O&D ONLINE", "EVOL T/P (O&D)", "CA", "R/R", "O&D", "R/R",
            "CA", "R/R", "O&D", "R/R", "T/P CA", "T/P CA ONLINE", "EVOL T/P (CA)", "T/P O&D",
            "T/P O&D ONLINE", "EVOL T/P (O&D)"
        ]

        # Define styles for headers, subtotals, top labels, date information, and borders
        header_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")  # Purple fill
        header_font = Font(bold=True, color="FFFFFF")  # White font
        header_alignment = Alignment(horizontal="center", vertical="center")

        subtotal_fill = PatternFill(start_color="5c4292", end_color="5c4292",
                                    fill_type="solid")  # Red fill for subtotals
        subtotal_font = Font(bold=True, color="FFFFFF")  # White font for subtotals, bold
        subtotal_alignment = Alignment(horizontal="center", vertical="center")
        subtotal_value_font = Font(bold=True, color="000000")  # Black font for subtotal values, bold

        date_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")  # Purple fill
        date_font = Font(bold=True, color="FFFFFF")  # White font for date labels
        date_alignment = Alignment(horizontal="center", vertical="center")

        top_label_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")  # Purple fill
        top_label_font = Font(bold=True, color="FFFFFF")  # White font for top labels
        top_label_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(border_style="thin", color="262626"),
            right=Side(border_style="thin", color="262626"),
            top=Side(border_style="thin", color="262626"),
            bottom=Side(border_style="thin", color="262626")
        )

        subtotals = ["SOUS TOTAL IAC", "SOUS TOTAL IHAC", "SOUS TOTAL MAC", "SOUS TOTAL MHAC"]

        # Get the current date information
        current_date = datetime.now()
        current_year = current_date.year
        last_month = current_date.month - 1 if current_date.month > 1 else 12
        last_month_name = datetime(1900, last_month, 1).strftime('%B')
        if last_month == 12:
            current_year -= 1  # Adjust the year if the last month is December from the previous year

        def write_dataframe_to_sheet(df, start_row, start_col, subtotal_text):
            # Add two rows above headers for labels
            # Last month label
            sheet.merge_cells(f"H{start_row}:U{start_row}")
            last_month_label_cell = sheet["H" + str(start_row)]
            last_month_label_cell.value = f"{last_month_name} {current_year}"
            last_month_label_cell.fill = date_fill
            last_month_label_cell.font = date_font
            last_month_label_cell.alignment = date_alignment
            last_month_label_cell.border = thin_border

            # Cumulative label from V to AI
            sheet.merge_cells(f"V{start_row}:AI{start_row}")
            cumulative_label_cell = sheet["V" + str(start_row)]
            cumulative_label_cell.value = f"January {current_year} to {last_month_name} {current_year}"
            cumulative_label_cell.fill = date_fill
            cumulative_label_cell.font = date_font
            cumulative_label_cell.alignment = date_alignment
            cumulative_label_cell.border = thin_border

            # Write top labels above headers (one row below the date labels)
            label_ranges = [
                ("Industrie", "H", "K"),
                ("Groupe AF KL", "L", "O"),
                ("% GROUPE AF KL", "P", "U"),
                ("Industrie", "V", "Y"),
                ("Groupe AF KL", "Z", "AC"),
                ("% GROUPE AF KL", "AD", "AI")
            ]

            for label, start_col_letter, end_col_letter in label_ranges:
                sheet.merge_cells(f"{start_col_letter}{start_row + 1}:{end_col_letter}{start_row + 1}")
                top_label_cell = sheet[f"{start_col_letter}{start_row + 1}"]
                top_label_cell.value = label
                top_label_cell.fill = top_label_fill
                top_label_cell.font = top_label_font
                top_label_cell.alignment = top_label_alignment
                top_label_cell.border = thin_border

            # Write headers below the top labels
            for col_idx, header in enumerate(headers, start=start_col):
                cell = sheet.cell(row=start_row + 2, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data below headers
            for r_idx in range(len(df)):
                for c_idx in range(len(df.columns)):
                    cell = sheet.cell(row=start_row + r_idx + 3, column=start_col + c_idx)
                    cell.value = df.iloc[r_idx, c_idx]
                    cell.border = thin_border

            # Write subtotal row and format
            subtotal_row = start_row + len(df) + 2  # Directly after data rows (moved one row up)
            start_merge_col = get_column_letter(start_col)
            end_merge_col = get_column_letter(start_col + 6)  # Merge columns A to G
            sheet.merge_cells(f"{start_merge_col}{subtotal_row}:{end_merge_col}{subtotal_row}")
            subtotal_cell = sheet.cell(row=subtotal_row, column=start_col)
            subtotal_cell.value = subtotal_text
            subtotal_cell.fill = subtotal_fill
            subtotal_cell.font = subtotal_font
            subtotal_cell.alignment = subtotal_alignment
            subtotal_cell.border = thin_border

            # Apply subtotal styling to all columns in the subtotal row after column G
            for col_idx in range(start_col + 7, start_col + len(headers)):
                cell = sheet.cell(row=subtotal_row, column=col_idx)
                cell.font = subtotal_value_font
                cell.border = thin_border

    # Write the data to the worksheet
    start_row = 7
    start_col = 1

    # Write IAC, IHAC, MAC, MHAC data
    for df_filtered, subtotal_text in zip(filtered_dfs, subtotals):
        write_dataframe_to_sheet(df_filtered, start_row, start_col, subtotal_text)
        # Update start_row for the next dataset
        start_row += len(df_filtered) + 4

    # Save the updated workbook
    updated_excel_path = 'csv/OK/excel_with_flight.xlsx'
    workbook.save(updated_excel_path)

    return updated_excel_path

def create_excel_train():
    # Load the CSV files into DataFrames
    euro_df = pd.read_csv('csv/OK/total_euro.csv')
    metro_df = pd.read_csv('csv/OK/total_metro.csv')

    # Drop the "O&D" column and columns ending with "_N_1"
    columns_to_exclude = ["O&D"] + [col for col in euro_df.columns if col.endswith("_N_1")]

    euro_df.drop(columns=columns_to_exclude, inplace=True, errors='ignore')
    metro_df.drop(columns=columns_to_exclude, inplace=True, errors='ignore')

    # Drop rows where "CA_jan_to_last_month_N" or "OD_jan_to_last_month_N" are empty (NaN)
    euro_df.dropna(subset=["CA_jan_to_last_month_N", "OD_jan_to_last_month_N"], inplace=True)
    metro_df.dropna(subset=["CA_jan_to_last_month_N", "OD_jan_to_last_month_N"], inplace=True)


    # Reorder columns as specified
    columns_order = [
        "CODE ORIGINE", "CODE DESTINATION", "LIBELLE ORIGINE", "LIBELLE DESTINATION",
        "CA_last_month_N", "RR_CA", "OD_last_month_N", "RR_OD",
        "CA_jan_to_last_month_N", "cumul_RR_CA", "OD_jan_to_last_month_N", "cumul_RR_OD"
    ]

    euro_df = euro_df[columns_order]
    metro_df = metro_df[columns_order]

    # Rename columns as specified
    rename_columns = {
        "CA_last_month_N": "CA",
        "CA_jan_to_last_month_N": "cumul_CA",
        "OD_last_month_N": "OD",
        "OD_jan_to_last_month_N": "cumul_OD"
    }

    euro_df.rename(columns=rename_columns, inplace=True)
    metro_df.rename(columns=rename_columns, inplace=True)

    # Load the existing Excel file
    excel_path = "csv/OK/excel_with_flight.xlsx"
    workbook = load_workbook(excel_path)

    # Select the worksheets "Etat 1.2" and "Etat 1.3"
    euro_sheet_name = "Etat 1.2"
    metro_sheet_name = "Etat 1.3"

    euro_sheet = workbook[euro_sheet_name] if euro_sheet_name in workbook.sheetnames else None
    metro_sheet = workbook[metro_sheet_name] if metro_sheet_name in workbook.sheetnames else None

    if not euro_sheet or not metro_sheet:
        print("One of the specified sheets does not exist in the workbook.")
        return

    # Define headers to be used for each table
    headers = list(euro_df.columns)

    # Define styles for headers, date information, and borders
    header_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")  # Purple fill
    header_font = Font(bold=True, color="FFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center")

    date_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")  # Purple fill
    date_font = Font(bold=True, color="FFFFFF")  # White font for date labels
    date_alignment = Alignment(horizontal="center", vertical="center")

    subtotal_fill = PatternFill(start_color="5c4292", end_color="5c4292",
                                fill_type="solid")  # Purple fill for subtotals
    subtotal_font = Font(bold=True, color="FFFFFF")  # White font for subtotals
    subtotal_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(border_style="thin", color="262626"),
        right=Side(border_style="thin", color="262626"),
        top=Side(border_style="thin", color="262626"),
        bottom=Side(border_style="thin", color="262626")
    )

    # Get the current date information
    current_date = datetime.now()
    current_year = current_date.year
    last_month = current_date.month - 1 if current_date.month > 1 else 12
    last_month_name = datetime(1900, last_month, 1).strftime('%B')
    if last_month == 12:
        current_year -= 1  # Adjust the year if the last month is December from the previous year

    def write_dataframe_to_sheet(df, sheet, start_row, start_col, global_text):
        # Unmerge all cells in the sheet to handle any existing merged cells
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))

        for col_idx in range(5, 9):  # E to H for "Last month and year"
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.border = thin_border

        for col_idx in range(9, 13):  # I to L for "Cumul de janvier à {last month} {current year}"
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.border = thin_border

        # Write the date labels above the top labels
        # "Last month and year" for the left part of the header
        sheet.merge_cells(f"E{start_row}:H{start_row}")
        last_month_label_cell = sheet["E" + str(start_row)]
        last_month_label_cell.value = f"{last_month_name} {current_year}"
        last_month_label_cell.fill = date_fill
        last_month_label_cell.font = date_font
        last_month_label_cell.alignment = date_alignment
        last_month_label_cell.border = thin_border

        # "Cumul de janvier à {last month} {current year}" for the right part of the header
        sheet.merge_cells(f"I{start_row}:L{start_row}")
        cumulative_label_cell = sheet["I" + str(start_row)]
        cumulative_label_cell.value = f"January {current_year} to {last_month_name} {current_year}"
        cumulative_label_cell.fill = date_fill
        cumulative_label_cell.font = date_font
        cumulative_label_cell.alignment = date_alignment
        cumulative_label_cell.border = thin_border

        # Write headers below the date labels
        for col_idx, header in enumerate(headers, start=start_col):
            cell = sheet.cell(row=start_row + 1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data below headers
        for r_idx in range(len(df)):
            for c_idx in range(len(df.columns)):
                cell = sheet.cell(row=start_row + r_idx + 2, column=start_col + c_idx)
                cell.value = df.iloc[r_idx, c_idx]
                cell.border = thin_border

        # Write global total row
        global_row = start_row + len(df) + 1  # Adjust row to move up by 1
        sheet.merge_cells(f"A{global_row}:D{global_row}")
        global_total_cell = sheet.cell(row=global_row, column=start_col)
        global_total_cell.value = global_text
        global_total_cell.fill = subtotal_fill
        global_total_cell.font = subtotal_font
        global_total_cell.alignment = subtotal_alignment
        global_total_cell.border = thin_border
        for col in range(start_col, start_col + len(headers)):
            cell = sheet.cell(row=global_row, column=col)
            cell.border = thin_border

        for col_idx in range(start_col + 4, start_col + len(headers)):
            cell = sheet.cell(row=global_row, column=col_idx)
            cell.font = Font(bold=True)  # Make the values bold
            cell.border = thin_border

    # Write euro data to "Etat 1.2"
    write_dataframe_to_sheet(metro_df, euro_sheet, start_row=4, start_col=1, global_text="GLOBAL RAIL")

    # Write metro data to "Etat 1.3"
    write_dataframe_to_sheet(euro_df, metro_sheet, start_row=4, start_col=1, global_text="GLOBAL RAIL")

    # Save the updated workbook
    updated_excel_path = 'csv/OK/excel_with_all.xlsx'
    workbook.save(updated_excel_path)

    # Return the path of the updated file
    return updated_excel_path

def extract_global():

    # Load the uploaded CSV file into a DataFrame
    file_path = 'csv/OK/total_flight.csv'
    df = pd.read_csv(file_path)

    # Extract the last row of the DataFrame
    last_row = df.tail(1)

    columns_to_keep = [
        'ORIGIN_AREA',
        'current_CA_INDUSTRIE_N', 'current_RR_CA_INDUSTRIE_N', 'current_OD_INDUSTRIE_N', 'current_RR_OD_INDUSTRIE_N',
        'current_CA_AFKL_N', 'current_RR_CA_AFKL_N', 'current_OD_AFKL_N', 'current_RR_OD_AFKL_N',
        'current_TP_CA_AFKL_N', 'current_TP_CA_AFKL_ONLINE_N', 'current_EVOL_TP_CA',
        'current_TP_OD_AFKL_N', 'current_TP_OD_AFKL_ONLINE_N', 'current_EVOL_TP_OD',
        'cumul_CA_INDUSTRIE_N', 'cumul_RR_CA_INDUSTRIE_N', 'cumul_OD_INDUSTRIE_N', 'cumul_RR_OD_INDUSTRIE_N',
        'cumul_CA_AFKL_N', 'cumul_RR_CA_AFKL_N', 'cumul_OD_AFKL_N', 'cumul_RR_OD_AFKL_N',
        'cumul_TP_CA_AFKL_N', 'cumul_TP_CA_AFKL_ONLINE_N', 'cumul_EVOL_TP_CA',
        'cumul_TP_OD_AFKL_N', 'cumul_TP_OD_AFKL_ONLINE_N', 'cumul_EVOL_TP_OD'
    ]

    # Keep only the specified columns in the last row extract
    filtered_last_row = last_row[columns_to_keep]
    filtered_last_row = filtered_last_row.round(2)

    # Save the last row to a new CSV file
    output_path = 'csv/OK/total_global.csv'
    filtered_last_row.to_csv(output_path, index=False)


def add_global(name_orga):
    import openpyxl
    from openpyxl.styles import Alignment
    import pandas as pd


    # Load the provided Excel and CSV files
    excel_path = "csv/OK/excel_with_all.xlsx"
    csv_path = f"csv/OK/total_global.csv"

    # Load the CSV file to get the "Total GLOBAL" row
    csv_data = pd.read_csv(csv_path)

    # Load the Excel workbook and the specific sheet
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook["Etat 1.1"]

    # Find the row index of "SOUS TOTAL MHAC" in the Excel sheet
    sous_total_mhac_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        if row[0] == "SOUS TOTAL MHAC":
            sous_total_mhac_row = idx
            break

    # Insert "Total GLOBAL" two lines after "SOUS TOTAL MHAC"
    if sous_total_mhac_row is not None and not csv_data.empty:
        # Get the row to append and reset its index
        total_global_row = csv_data.reset_index(drop=True).iloc[0]

        # Prepare the insertion row number
        insert_row = sous_total_mhac_row + 2

        # Merge columns A to G for the "Total GLOBAL" label
        sheet.merge_cells(start_row=insert_row, start_column=1, end_row=insert_row, end_column=7)
        cell = sheet.cell(row=insert_row, column=1)
        cell.value = "Total GLOBAL"
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write the rest of the values starting from column H
        # Writing all columns starting from H without changing their order
        for col_index, value in enumerate(total_global_row.iloc[1:], start=8):  # Start at column H, skip "ANNEX_C"
            sheet.cell(row=insert_row, column=col_index, value=value)

        # Save the updated workbook
        output_excel_path = f"csv/OK/{name_orga}_{last_month_name}_{current_year}.xlsx"
        workbook.save(output_excel_path)

    else:
        print("Total GLOBAL row not found in the CSV file or 'SOUS TOTAL MHAC' not found in the Excel sheet.")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def joli(name_orga):
    # Load the provided Excel workbook
    file_path = f"csv/OK/{name_orga}_{last_month_name}_{current_year}.xlsx"
    workbook = openpyxl.load_workbook(file_path)

    # Define the white fill style for the added columns
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Define the styles for "TOTAL GLOBAL"
    header_fill = PatternFill(start_color="5C4292", end_color="5C4292", fill_type="solid")  # Purple fill with better contrast
    header_font_white = Font(bold=True, color="FFFFFF")  # White font for header text
    header_font_black = Font(bold=True, color="000000")  # Black font for other cells in the row
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define border style for the row "Total GLOBAL"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Update the sheet "Etat 1.1"
    sheet = workbook["Etat 1.1"]

    # Iterate through merged cells to find "Total GLOBAL" and apply styles
    for merged_cell in sheet.merged_cells.ranges:
        top_left_cell = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col)
        if top_left_cell.value and "Total GLOBAL" in str(top_left_cell.value):
            # Apply styles to the merged cells (A to G) for "Total GLOBAL"
            for row in sheet.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                                       min_col=merged_cell.min_col, max_col=merged_cell.max_col):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font_white
                    cell.alignment = header_alignment

            # Apply bold style to the entire row of "Total GLOBAL" (from A to the last column), with black text for columns after G
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=merged_cell.min_row, column=col)
                if col <= 7:  # Columns A to G should have white text
                    cell.font = header_font_white
                else:  # Other columns should have black text
                    cell.font = header_font_black

            # Apply border to the specified range of cells (A to U and W to AJ)
            for col in range(1, 22):  # Columns A to U (1 to 20)
                cell = sheet.cell(row=merged_cell.min_row, column=col)
                cell.border = thin_border

            for col in range(22, 36):  # Columns W to AJ (23 to 36)
                cell = sheet.cell(row=merged_cell.min_row, column=col)
                cell.border = thin_border

    # Create an empty column at "V" and shift contents to the right without breaking merged cells
    merged_cells = list(sheet.merged_cells.ranges)
    sheet.insert_cols(22)
    new_column_letter = get_column_letter(22)
    sheet.column_dimensions[new_column_letter].width = 0.56  # Set width to 2 mm
    for row in range(1, sheet.max_row + 1):
        sheet[f'{new_column_letter}{row}'].fill = white_fill

    # Update merged cell ranges to account for the inserted column
    sheet.merged_cells.ranges = []  # Clear existing merges
    for merged_cell in merged_cells:
        if merged_cell.min_col >= 22:
            new_min_col = merged_cell.min_col + 1
            new_max_col = merged_cell.max_col + 1
        else:
            new_min_col = merged_cell.min_col
            new_max_col = merged_cell.max_col
        new_range = f"{get_column_letter(new_min_col)}{merged_cell.min_row}:{get_column_letter(new_max_col)}{merged_cell.max_row}"
        sheet.merge_cells(new_range)

    # Update the sheets "Etat 1.2" and "Etat 1.3" by adding a column at "I"
    sheets_to_update = ["Etat 1.2", "Etat 1.3"]

    for sheet_name in sheets_to_update:
        sheet = workbook[sheet_name]
        merged_cells = list(sheet.merged_cells.ranges)
        sheet.insert_cols(9)
        new_column_letter = get_column_letter(9)
        sheet.column_dimensions[new_column_letter].width = 0.56  # Set width to 2 mm
        for row in range(1, sheet.max_row + 1):
            sheet[f'{new_column_letter}{row}'].fill = white_fill
        sheet.merged_cells.ranges = []  # Clear existing merges
        for merged_cell in merged_cells:
            if merged_cell.min_col >= 9:
                new_min_col = merged_cell.min_col + 1
                new_max_col = merged_cell.max_col + 1
            else:
                new_min_col = merged_cell.min_col
                new_max_col = merged_cell.max_col
            new_range = f"{get_column_letter(new_min_col)}{merged_cell.min_row}:{get_column_letter(new_max_col)}{merged_cell.max_row}"
            sheet.merge_cells(new_range)

    # Convert specified columns to percentages in "Etat 1.1" and format to 2 decimal places
    sheet = workbook["Etat 1.1"]
    columns_to_convert_etat_1_1 = ["I", "K", "M", "O", "P", "Q", "S", "T", "X", "Z", "AB", "AD", "AE", "AF", "AH", "AI"]

    for col_letter in columns_to_convert_etat_1_1:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        for row in range(4, sheet.max_row + 1):  # Starting from row 4 (after headers)
            cell = sheet.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / 100  # Convert to percentage
                cell.number_format = '0.00%'  # Set cell format to percentage with 2 decimals

    # Convert specified columns to percentages in "Etat 1.2" and "Etat 1.3" and format to 2 decimal places
    columns_to_convert_etat_1_2_1_3 = ["F", "H", "K", "M"]
    for sheet_name in ["Etat 1.2", "Etat 1.3"]:
        sheet = workbook[sheet_name]
        for col_letter in columns_to_convert_etat_1_2_1_3:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            for row in range(4, sheet.max_row + 1):  # Starting from row 4 (after headers)
                cell = sheet.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100  # Convert to percentage
                    cell.number_format = '0.00%'  # Set cell format to percentage with 2 decimals

    # Save the updated workbook
    output_path = f"csv/OK/{name_orga}_{last_month_name}_{current_year}.xlsx"
    workbook.save(output_path)





